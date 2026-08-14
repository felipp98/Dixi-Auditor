import os
import sys

# Garante que a raiz do projeto esteja no sys.path para resolução do pacote 'src'
_PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

import ctypes
import re
import json

# Habilita suporte a High DPI (DPI Awareness) no Windows para telas de 800x600 até 4K (3840x2160)
if sys.platform.startswith("win"):
    try:
        ctypes.windll.shcore.SetProcessDpiAwareness(2)
    except Exception:
        try:
            ctypes.windll.user32.SetProcessDPIAware()
        except Exception:
            pass


import io

import logging

import threading

import keyring

import calendar

from datetime import datetime

from typing import List, Dict, Optional, Tuple

from dataclasses import dataclass



import pandas as pd

import requests

from openpyxl import load_workbook

from openpyxl.styles import PatternFill, Font

import tkinter as tk

from tkinter import messagebox, ttk, filedialog

from PIL import Image, ImageTk

from playwright.sync_api import sync_playwright



try:
    from src.services.justificativa_service import gerar_pdf_justificativa, enviar_email_smtp, normalizar_dia
    from src.utils.formatters import formatar_mes_competencia, obter_mes_extenso
    from src.services.autentique_service import enviar_justificativa_autentique, listar_documentos_autentique, resgatar_documento_autentique
except ImportError:
    from justificativa_service import gerar_pdf_justificativa, formatar_mes_competencia, obter_mes_extenso, enviar_email_smtp
    from autentique_service import enviar_justificativa_autentique, listar_documentos_autentique, resgatar_documento_autentique



# Configura o Logging para salvar erros em arquivo local

log_file = os.path.join(os.path.expanduser("~"), "dixi_auditor.log")

logging.basicConfig(

    filename=log_file,

    level=logging.ERROR,

    format="%(asctime)s - %(levelname)s - %(message)s"

)



# --- MODELOS DE DADOS ---

@dataclass

class MarcacaoDia:

    data_id: str  

    data_formatada: str

    segundos_trabalhados: int

    saldo_segundos: int

    is_pendencia: bool

    horarios: List[str]

    obs: str = ""



# --- SERVIÇO DE API ---

class DixiService:

    def __init__(self):

        self.base_url = "https://webapiponto.dixiponto.com.br:8899"

        self.session = requests.Session()

        self.token: Optional[str] = None

        self.user_id: Optional[int] = None

        self.user_name: Optional[str] = None

        self.user_cargo: Optional[str] = None



    def authenticate(self, user: str, password: str) -> bool:

        try:

            resp = self.session.post(f"{self.base_url}/login_", json={

                "usuario": user, "senha": password, "unidade": "pagare", "suporte": False

            }, timeout=15)

            resp.raise_for_status()

            data = resp.json()

            if data.get("success"):

                self.token = data["data"]["token"]

                user_obj = data["data"].get("usuario", {})

                func_info = user_obj.get("funcionario", {})

                self.user_id = func_info.get("idFuncionario")

                self.user_name = func_info.get("nomeFuncionario") or func_info.get("nome") or user_obj.get("nome") or user

                self.user_cargo = func_info.get("descricaoCargo") or func_info.get("cargo") or "Colaborador"

                self.user_email = func_info.get("email") or user_obj.get("email") or user

                self.session.headers.update({"Authorization": f"bearer {self.token}"})

                # Guarda as credenciais com segurança no Windows

                keyring.set_password("DixiPontoApp", "last_user", user)

                keyring.set_password("DixiPontoApp", user, password)

                return True

            return False

        except Exception as e:

            logging.error(f"Erro na autenticação: {e}")

            return False



    def fetch_history(self, start: str, end: str) -> List[Dict]:

        params = {"dataInicial": start, "dataFinal": end, "idRegistro": self.user_id}

        resp = self.session.get(f"{self.base_url}/self_/historicoPonto", params=params, timeout=20)

        resp.raise_for_status()

        return resp.json().get("data", [])



# --- MECANISMO DE CÁLCULO ---

class PontoEngine:

    JORNADA_SEG = 8 * 3600

    TOLERANCIA_SEG = 600

    MIN_ALMOCO_SEG = 3600  # 1 hora de almoço (3600s)



    @classmethod

    def process_horarios(cls, raw_horarios: List[str], data_id: str, data_formatada: str, obs: str = "") -> MarcacaoDia:

        raw_horarios = sorted(raw_horarios)

        qtd_batidas = len(raw_horarios)

        total_sec = 0

        

        for i in range(0, qtd_batidas // 2 * 2, 2):

            h1 = datetime.strptime(raw_horarios[i], "%H:%M")

            h2 = datetime.strptime(raw_horarios[i+1], "%H:%M")

            diff_sec = (h2 - h1).total_seconds()

            if diff_sec < 0:

                diff_sec += 24 * 3600  # Suporte para virada de dia/turno noturno

            total_sec += int(diff_sec)



        # Regra de Almoço: Se voltou antes de 1 hora (3600s), desconsidera o tempo antecipado como hora extra

        if qtd_batidas >= 4:

            s1 = datetime.strptime(raw_horarios[1], "%H:%M")

            e2 = datetime.strptime(raw_horarios[2], "%H:%M")

            intervalo_almoco = (e2 - s1).total_seconds()

            if intervalo_almoco < 0:

                intervalo_almoco += 24 * 3600

            

            if intervalo_almoco < cls.MIN_ALMOCO_SEG:

                desconto_antecipacao = cls.MIN_ALMOCO_SEG - intervalo_almoco

                total_sec -= int(desconto_antecipacao)



        is_pendencia = (qtd_batidas % 2 != 0) or (qtd_batidas == 2)

        

        saldo = 0
        if total_sec > 0:
            diff = total_sec - cls.JORNADA_SEG
            if diff > 0:
                # Crédito (Horas Extras): computa 100% integralmente sem desconto
                saldo = diff
            elif diff < 0:
                # Débito (Atrasos): aplica o perdão/tolerância de 10 min se exceder 10 min
                if abs(diff) > cls.TOLERANCIA_SEG:
                    saldo = diff + cls.TOLERANCIA_SEG
                else:
                    saldo = diff



        return MarcacaoDia(

            data_id=data_id,

            data_formatada=data_formatada,

            segundos_trabalhados=total_sec,

            saldo_segundos=saldo,

            is_pendencia=is_pendencia,

            horarios=raw_horarios,

            obs=obs

        )



    @classmethod

    def process_day(cls, day_data: Dict) -> MarcacaoDia:

        raw_horarios = sorted([m["hora"] for m in day_data["marcacoes"]])

        dt_obj = datetime.strptime(str(day_data["data"]), "%Y%m%d")

        return cls.process_horarios(raw_horarios, str(day_data["data"]), dt_obj.strftime("%d/%m/%Y"))



# --- EXPORTADOR EXCEL ---

class ExcelExporter:

    @staticmethod

    def format_time(seconds: int, show_sign: bool = False) -> str:

        abs_sec = abs(seconds)

        h, m = divmod(abs_sec // 60, 60)

        sign = ("+" if seconds > 0 else "-") if show_sign and seconds != 0 else ""

        return f"{sign}{int(h):02d}:{int(m):02d}"



    def generate(self, data: List[MarcacaoDia], path: str, ignore_today: bool = True):

        rows = []

        sum_saldo = 0 

        today_str = datetime.now().strftime("%d/%m/%Y")



        max_horarios = max([len(m.horarios) for m in data]) if data else 0

        max_cols = max(6, max_horarios)

        if max_cols % 2 != 0:

            max_cols += 1



        for m in data:

            is_today = (m.data_formatada == today_str) and ignore_today

            if not is_today:

                sum_saldo += m.saldo_segundos



            punches = (m.horarios + [""] * max_cols)[:max_cols]

            saldo_str = "00:00" if is_today else self.format_time(m.saldo_segundos, True)

            obs_str = m.obs if m.obs else ("EM ANDAMENTO" if is_today else ("FALTA BATIDA" if m.is_pendencia else ""))

            

            row = [m.data_formatada] + punches + [

                self.format_time(m.segundos_trabalhados),

                saldo_str,

                obs_str

            ]

            rows.append(row)

        

        headers = ["Data"]

        for i in range(1, (max_cols // 2) + 1):

            headers.extend([f"E{i}", f"S{i}"])

        headers.extend(["Total", "Saldo", "Obs"])

        

        df = pd.DataFrame(rows, columns=headers)

        df.to_excel(path, index=False)

        self._finalize_excel(path, sum_saldo, max_cols)



    def _finalize_excel(self, path, total_s, max_cols):

        wb = load_workbook(path)

        ws = wb.active

        last_row = ws.max_row + 1

        

        from openpyxl.utils import get_column_letter

        col_saldo_letter = get_column_letter(3 + max_cols)

        col_obs_letter = get_column_letter(4 + max_cols)

        

        ws[f"A{last_row}"] = "SALDO ACUMULADO"

        ws[f"{col_saldo_letter}{last_row}"] = self.format_time(total_s, True)

        

        fill_total = PatternFill(start_color="DDEBF7", fill_type="solid")

        fill_red = PatternFill(start_color="FFC7CE", fill_type="solid")

        fill_green = PatternFill(start_color="C6EFCE", fill_type="solid")

        fill_yellow = PatternFill(start_color="FFF2CC", fill_type="solid")

        fill_olive = PatternFill(start_color="E2EFDA", fill_type="solid")

        bold_font = Font(bold=True)



        for cell in ws[last_row]:

            cell.fill = fill_total

            cell.font = bold_font



        for row in range(2, ws.max_row):

            saldo_cell = ws[f"{col_saldo_letter}{row}"]

            obs_cell = ws[f"{col_obs_letter}{row}"]

            val = str(saldo_cell.value)

            if "+" in val: 

                saldo_cell.fill = fill_green

            elif "-" in val and val != "00:00": 

                saldo_cell.fill = fill_red

            if obs_cell.value == "FALTA BATIDA": 

                obs_cell.fill = fill_red

            elif obs_cell.value == "EM ANDAMENTO":

                obs_cell.fill = fill_olive

            elif obs_cell.value and any(k in str(obs_cell.value) for k in ["IA", "Ajust", "Abon"]):

                obs_cell.fill = fill_yellow



        wb.save(path)



# --- SERVIÇO DE ANÁLISE POR IA ---

class IAAnalistaPonto:

    @staticmethod

    def get_api_key() -> str:

        key = os.environ.get("ANTHROPIC_AUTH_TOKEN") or os.environ.get("OPENROUTER_API_KEY")

        if not key:

            key = keyring.get_password("DixiPontoApp", "openrouter_token") or ""

        return key



    @staticmethod

    def encontrar_sugestoes_duplicadas(data: List[MarcacaoDia]) -> Tuple[List[Dict], str]:

        sugestoes_ajustes = []

        texto_sugestoes = ""

        for m in data:

            if not m.horarios or len(m.horarios) < 2:

                continue

            if len(m.horarios) != len(set(m.horarios)):

                h_unicos = sorted(list(dict.fromkeys(m.horarios)))

                sugestoes_ajustes.append({

                    "data": m.data_formatada,

                    "horarios": h_unicos,

                    "obs": "Corrigido via IA: duplicação de batida removida"

                })

                texto_sugestoes += f"  • Dia {m.data_formatada}: Batida duplicada ({', '.join(m.horarios)}) -> Sugestão: remover duplicidade ({', '.join(h_unicos)})\n"

            else:

                cleaned = [m.horarios[0]]

                has_near_dup = False

                for h in m.horarios[1:]:

                    try:

                        t_prev = datetime.strptime(cleaned[-1], "%H:%M")

                        t_curr = datetime.strptime(h, "%H:%M")

                        diff_min = abs((t_curr - t_prev).total_seconds()) / 60

                        if diff_min <= 2:

                            has_near_dup = True

                        else:

                            cleaned.append(h)

                    except Exception:

                        cleaned.append(h)

                if has_near_dup:

                    sugestoes_ajustes.append({

                        "data": m.data_formatada,

                        "horarios": cleaned,

                        "obs": "Corrigido via IA: batidas de curto intervalo unificadas"

                    })

                    texto_sugestoes += f"  • Dia {m.data_formatada}: Batidas muito próximas ({', '.join(m.horarios)}) -> Sugestão: unificar ({', '.join(cleaned)})\n"

        return sugestoes_ajustes, texto_sugestoes



    @classmethod

    def analisar_ponto(cls, data: List[MarcacaoDia], api_key: Optional[str] = None, instrucoes_usuario: Optional[str] = None, ignore_today: bool = True) -> Tuple[str, List[Dict]]:

        token = api_key or cls.get_api_key()

        today_str = datetime.now().strftime("%d/%m/%Y")

        

        total_dias = len(data)

        

        if ignore_today:

            dias_pendencia = [m for m in data if m.is_pendencia and m.data_formatada != today_str]

            dias_extras = [m for m in data if m.saldo_segundos > 0 and m.data_formatada != today_str]

            dias_atraso = [m for m in data if m.saldo_segundos < 0 and m.data_formatada != today_str]

            saldo_total = sum(m.saldo_segundos for m in data if m.data_formatada != today_str)

        else:

            dias_pendencia = [m for m in data if m.is_pendencia]

            dias_extras = [m for m in data if m.saldo_segundos > 0]

            dias_atraso = [m for m in data if m.saldo_segundos < 0]

            saldo_total = sum(m.saldo_segundos for m in data)



        saldo_str = ExcelExporter.format_time(saldo_total, True)

        

        resumo_dados = f"Total de dias auditados: {total_dias}\n"

        resumo_dados += f"Saldo acumulado no período (desconsiderando dia atual em andamento): {saldo_str}\n" if ignore_today else f"Saldo acumulado no período: {saldo_str}\n"

        resumo_dados += f"Dias com batida pendente/faltante: {len(dias_pendencia)}\n"

        resumo_dados += f"Dias com saldo positivo (HE): {len(dias_extras)}\n"

        resumo_dados += f"Dias com saldo negativo (atrasos): {len(dias_atraso)}\n\n"

        resumo_dados += "Detalhamento por dia:\n"

        

        for m in data[:31]:

            is_today = (m.data_formatada == today_str) and ignore_today

            if is_today:

                resumo_dados += f"- Data: {m.data_formatada} | Batidas: {', '.join(m.horarios)} | Trabalhado: {ExcelExporter.format_time(m.segundos_trabalhados)} | (EM ANDAMENTO - IGNORADO NO SALDO)\n"

            else:

                resumo_dados += f"- Data: {m.data_formatada} | Batidas: {', '.join(m.horarios)} | Trabalhado: {ExcelExporter.format_time(m.segundos_trabalhados)} | Saldo: {ExcelExporter.format_time(m.saldo_segundos, True)} {'(PENDÊNCIA)' if m.is_pendencia else ''}\n"



        parsed_ajustes = []



        if token:

            try:

                base_url = os.environ.get("ANTHROPIC_BASE_URL", "https://openrouter.ai/api")

                endpoint = f"{base_url.rstrip('/')}/v1/chat/completions"

                model = os.environ.get("ANTHROPIC_DEFAULT_SONNET_MODEL", "meta-llama/llama-3.3-70b-instruct:free")

                

                headers = {

                    "Authorization": f"Bearer {token}",

                    "Content-Type": "application/json"

                }

                prompt = (

                    "Você é um especialista em auditoria de RH e cartão de ponto.\n"

                    "Analise a jornada abaixo de forma clara, profissional e objetiva em português:\n"

                    "1. Faça um resumo geral do período auditado.\n"

                    "2. Destaque inconsistências graves (batidas ímpares, faltas ou atrasos recorrentes).\n"

                    "3. Dê recomendações para o gestor ou funcionário.\n\n"

                    f"DADOS DO PONTO:\n{resumo_dados}"

                )

                if instrucoes_usuario:

                    prompt += (

                        "\n\nSOLICITAÇÃO DE REAJUSTE / INSTRUÇÃO DO USUÁRIO PARA RECALCULAR:\n"

                        f"\"{instrucoes_usuario}\"\n\n"

                        "Por favor, considere as edições/ajustes solicitados acima pelo usuário nos pontos citados, recalculando a análise, saldos e considerações com base nessas instruções.\n"

                        "IMPORTANTE: Se a instrução do usuário definir horários específicos (ex: 'saída às 18:00') ou solicitar abono de faltas/pendências, inclua AO FINAL da resposta um bloco JSON estruturado exatamente assim:\n"

                        "```json\n"

                        "{\n"

                        "  \"ajustes\": [\n"

                        "    {\"data\": \"DD/MM/YYYY\", \"horarios\": [\"08:00\", \"12:00\", \"13:00\", \"18:00\"], \"obs\": \"Ajustado via IA: saída 18:00\"},\n"

                        "    {\"data\": \"DD/MM/YYYY\", \"abono\": true, \"obs\": \"Abonado via IA\"}\n"

                        "  ],\n"

                        "  \"enviar_justificativa\": true\n"

                        "}\n"

                        "```\n"

                        "Nota: Inclua \"enviar_justificativa\": true apenas se o usuário solicitar enviar, gerar ou encaminhar a justificativa/folha para o RH/Autentique."

                    )

                payload = {

                    "model": model,

                    "messages": [{"role": "user", "content": prompt}],

                    "temperature": 0.3

                }

                resp = requests.post(endpoint, json=payload, headers=headers, timeout=20)

                auto_enviar = False

                if resp.status_code == 200:

                    res_json = resp.json()

                    content = res_json["choices"][0]["message"]["content"]

                    json_match = re.search(r'```json\s*(\{.*?\})\s*```', content, re.DOTALL)

                    if json_match:

                        try:

                            data_json = json.loads(json_match.group(1))

                            parsed_ajustes = data_json.get("ajustes", [])

                            auto_enviar = bool(data_json.get("enviar_justificativa", False))

                            content = re.sub(r'```json\s*(\{.*?\})\s*```', '', content).strip()

                        except Exception as je:

                            logging.error(f"Erro ao parsear JSON de ajustes: {je}")

                    

                    # Checagem por palavra-chave se o usuário pediu envio

                    if instrucoes_usuario and any(k in instrucoes_usuario.lower() for k in ["enviar", "envie", "gerar", "justificativa", "rh", "autentique"]):

                        auto_enviar = True

                    return content, parsed_ajustes, auto_enviar

            except Exception as e:
                logging.error(f"Erro na API de IA: {e}")

        sugestoes_ajustes, texto_sugestoes = cls.encontrar_sugestoes_duplicadas(data)

        if instrucoes_usuario:
            user_msg_low = instrucoes_usuario.lower()
            if any(k in user_msg_low for k in ["sim", "aplicar", "corrigir", "aceitar", "sugest", "duplicad", "remover", "ok"]):
                if not parsed_ajustes and sugestoes_ajustes:
                    parsed_ajustes = sugestoes_ajustes

            for line in instrucoes_usuario.split('\n'):
                m_date = re.search(r'(\d{2}/\d{2}(?:/\d{4})?)', line)
                if m_date:
                    dt_str = m_date.group(1)
                    if len(dt_str) == 5 and data:
                        year = data[0].data_formatada.split('/')[-1]
                        dt_str = f"{dt_str}/{year}"
                    times = re.findall(r'\b\d{2}:\d{2}\b', line)
                    if times:
                        parsed_ajustes.append({"data": dt_str, "horarios": times, "obs": "Ajustado via IA (Local)"})
                    elif any(k in line.lower() for k in ["abon", "falt", "desconsider"]):
                        parsed_ajustes.append({"data": dt_str, "abono": True, "obs": "Abonado via IA (Local)"})

        fallback = "🤖 ANÁLISE AUTOMÁTICA DE AUDITORIA (Local)\n"
        fallback += "=" * 45 + "\n\n"
        fallback += f"• Período Auditado: {total_dias} dias registrados.\n"
        fallback += f"• Saldo Geral Acumulado: {saldo_str}\n\n"

        if texto_sugestoes:
            fallback += f"💡 SUGESTÃO DE CORREÇÃO AUTOMÁTICA DETECTADA:\n{texto_sugestoes}\n👉 Clique no botão '✨ Aplicar Sugestão' ou digite 'Sim' para aplicar automaticamente!\n\n"

        if instrucoes_usuario:
            fallback += f"✏️ AJUSTES E INSTRUÇÕES DIGITADAS:\n   \"{instrucoes_usuario}\"\n\n"
            if parsed_ajustes:
                fallback += f"✅ {len(parsed_ajustes)} ajuste(s) identificado(s) e aplicado(s) na tabela do aplicativo!\n\n"
        
        if dias_pendencia:
            fallback += f"⚠️ ATENÇÃO: Identificamos {len(dias_pendencia)} dia(s) com batidas pendentes/ímpares:\n"

            for d in dias_pendencia:

                fallback += f"   - Dia {d.data_formatada}: Marcações ({', '.join(d.horarios)})\n"

            fallback += "   -> Ação recomendada: Solicitar ajuste manual ou abono no sistema Dixi.\n\n"

        else:

            fallback += "✅ Nenhuma pendência de marcação ímpar detectada no período.\n\n"

            

        if len(dias_atraso) > 0:

            fallback += f"🔴 Atrasos Registrados: {len(dias_atraso)} dia(s) fecharam com saldo negativo.\n"

        if len(dias_extras) > 0:

            fallback += f"🟢 Horas Extras: {len(dias_extras)} dia(s) fecharam com saldo positivo.\n"



        auto_enviar = False

        if instrucoes_usuario and any(k in instrucoes_usuario.lower() for k in ["enviar", "envie", "gerar", "justificativa", "rh", "autentique"]):

            auto_enviar = True



        return fallback, parsed_ajustes, auto_enviar



# --- SELETOR DE DATAS CUSTOMIZADO ---

class DateSelector(ttk.Frame):

    def __init__(self, parent, default_day="01", default_month="07", default_year="2026"):

        super().__init__(parent)

        

        self.months = {

            "Janeiro": 1, "Fevereiro": 2, "Março": 3, "Abril": 4,

            "Maio": 5, "Junho": 6, "Julho": 7, "Agosto": 8,

            "Setembro": 9, "Outubro": 10, "Novembro": 11, "Dezembro": 12

        }

        self.month_names = list(self.months.keys())

        

        self.cb_day = ttk.Combobox(self, width=3, state="readonly")

        self.cb_day.pack(side="left", padx=2)

        

        ttk.Label(self, text="/", font=("Segoe UI", 10, "bold")).pack(side="left")

        

        self.cb_month = ttk.Combobox(self, values=self.month_names, width=10, state="readonly")

        month_num = int(default_month)

        month_name = [name for name, num in self.months.items() if num == month_num][0]

        self.cb_month.set(month_name)

        self.cb_month.pack(side="left", padx=2)

        

        ttk.Label(self, text="/", font=("Segoe UI", 10, "bold")).pack(side="left")

        

        self.ent_year = ttk.Entry(self, width=5)

        self.ent_year.insert(0, default_year)

        self.ent_year.pack(side="left", padx=2)



        self.cb_month.bind("<<ComboboxSelected>>", self.update_days)

        self.ent_year.bind("<FocusOut>", self.update_days)

        self.ent_year.bind("<KeyRelease>", self.update_days)



        self.update_days(default_day=default_day)



    def update_days(self, event=None, default_day=None):

        try:

            year = int(self.ent_year.get().strip())

        except ValueError:

            year = datetime.now().year

            

        month_name = self.cb_month.get()

        month_num = self.months.get(month_name, 7)

        

        _, max_days = calendar.monthrange(year, month_num)

        

        days_list = [f"{i:02d}" for i in range(1, max_days + 1)]

        self.cb_day["values"] = days_list

        

        curr_day = default_day if default_day else self.cb_day.get()

        if not curr_day:

            curr_day = "01"

            

        try:

            if int(curr_day) > max_days:

                curr_day = f"{max_days:02d}"

        except ValueError:

            curr_day = "01"

            

        self.cb_day.set(curr_day)



    def get_date(self) -> datetime:

        day = self.cb_day.get()

        month_num = self.months[self.cb_month.get()]

        year = self.ent_year.get().strip()

        

        try:

            return datetime.strptime(f"{day}/{month_num:02d}/{year}", "%d/%m/%Y")

        except ValueError:

            raise ValueError(f"Data inválida: {day}/{month_num:02d}/{year}")



# --- INTERFACE PRINCIPAL ---

class AppPonto(tk.Tk):

    def __init__(self):

        super().__init__()

        self.title("Dixi Auditor - Pagare")

        self.geometry("460x640")

        self.service = DixiService()

        self.exporter = ExcelExporter()

        self.processed_data: List[MarcacaoDia] = []

        

        # Configuração de Estilo Visual

        self.style = ttk.Style()

        self.style.theme_use("clam")

        

        self.bg_color = "#F8FAF9"

        self.surface_color = "#FFFFFF"

        self.primary_color = "#6ACC32"

        self.primary_dark = "#4C9A23"

        self.primary_soft = "#EAF5E3"

        self.text_color = "#1E293B"

        self.muted_text_color = "#64748B"

        self.border_color = "#080808"


        self.danger_color = "#C5221F"

        self.warning_bg = "#FEF7E0"

        self.logo_image = None

        self.logo_title_image = None

        self.logo_brand_image = None

        

        self.configure(bg=self.bg_color)

        self.style.configure("TFrame", background=self.bg_color)

        self.style.configure("TLabel", background=self.bg_color, foreground=self.text_color, font=("Segoe UI", 10))

        self.style.configure("TButton",

                             background=self.primary_color,

                             foreground="white",

                             font=("Segoe UI", 10, "bold"),

                             padding=6,

                             bordercolor=self.primary_color,

                             lightcolor=self.primary_color,

                             darkcolor=self.primary_color,

                             relief="flat")

        self.style.map("TButton",

                       background=[("active", self.primary_dark), ("disabled", "#BDC3C7")],

                       bordercolor=[("active", self.primary_dark), ("disabled", "#BDC3C7")],

                       lightcolor=[("active", self.primary_dark), ("disabled", "#BDC3C7")],

                       darkcolor=[("active", self.primary_dark), ("disabled", "#BDC3C7")])

        self.style.configure("TEntry",

                             fieldbackground=self.surface_color,

                             bordercolor="#BDC3C7",

                             lightcolor="#BDC3C7",

                             darkcolor="#BDC3C7")

        self.style.configure("TCombobox", fieldbackground=self.surface_color)



        self._load_brand_assets()

        

        self._init_login_ui()



    def _find_logo_path(self) -> Optional[str]:

        base_dir = os.path.dirname(os.path.abspath(__file__))

        candidates = [

            os.path.join(base_dir, "assets", "icons", "logo.svg"),

            os.path.join(os.path.dirname(base_dir), "assets", "icons", "logo.svg"),

            os.path.join(base_dir, "assets", "icons", "logo-pagare.svg"),

            os.path.join(os.path.dirname(base_dir), "assets", "icons", "logo-pagare.svg"),

            os.path.join(base_dir, "assets", "icons", "logo_pagare.svg"),

            os.path.join(os.path.dirname(base_dir), "assets", "icons", "logo_pagare.svg"),

        ]

        for path in candidates:

            if os.path.exists(path):

                return path

        return None



    def _load_brand_assets(self):

        logo_path = self._find_logo_path()

        if not logo_path:

            return



        try:

            with open(logo_path, "r", encoding="utf-8") as svg_file:

                svg_markup = svg_file.read()



            with sync_playwright() as p:

                browser = p.chromium.launch(headless=True)

                page = browser.new_page(viewport={"width": 1920, "height": 1080, "device_scale_factor": 2})

                page.set_content(

                    f"""

                    <html>

                        <body style="margin:0; background:transparent; display:flex; align-items:center; justify-content:center;">

                            {svg_markup}

                        </body>

                    </html>

                    """

                )

                svg_bytes = page.locator("svg").screenshot(omit_background=True)

                browser.close()



            logo = Image.open(io.BytesIO(svg_bytes)).convert("RGBA")

            symbol_logo = logo.crop((0, 280, 520, 860))

            self.logo_title_image = ImageTk.PhotoImage(symbol_logo.resize((28, 28), Image.LANCZOS))

            self.logo_brand_image = ImageTk.PhotoImage(logo.resize((100, 60), Image.LANCZOS))

            self.logo_image = self.logo_brand_image

            self.iconphoto(True, self.logo_title_image)

        except Exception as exc:

            logging.error(f"Erro ao carregar logo: {exc}")



    def _set_entry_focus_state(self, shell: tk.Frame, focused: bool):

        shell.configure(

            highlightbackground=self.primary_color if focused else self.border_color,

            highlightcolor=self.primary_color if focused else self.border_color,

            bg="#FFFFFF" if focused else "#F8FBF2"

        )



    def _build_login_input(self, parent: tk.Frame, label_text: str, show: Optional[str] = None) -> tk.Entry:

        field_group = tk.Frame(parent, bg=self.surface_color)

        field_group.pack(fill="x", pady=(0, 10))



        tk.Label(

            field_group,

            text=label_text,

            bg=self.surface_color,

            fg=self.text_color,

            font=("Segoe UI", 9, "bold")

        ).pack(anchor="w", pady=(0, 4))



        field_shell = tk.Frame(

            field_group,

            bg="#F8FBF2",

            highlightbackground=self.border_color,

            highlightcolor=self.primary_color,

            highlightthickness=1,

            bd=0

        )

        field_shell.pack(fill="x")



        entry = tk.Entry(

            field_shell,

            font=("Segoe UI", 11),

            bg="#F8FBF2",

            fg=self.text_color,

            insertbackground=self.text_color,

            relief="flat",

            bd=0,

            highlightthickness=0,

            show=show

        )

        entry.pack(fill="x", padx=12, pady=9)

        entry.bind("<FocusIn>", lambda _e: self._set_entry_focus_state(field_shell, True))

        entry.bind("<FocusOut>", lambda _e: self._set_entry_focus_state(field_shell, False))

        return entry



    def _bind_button_hover(self, button: tk.Button):

        button.bind("<Enter>", lambda _e: button.config(bg=self.primary_dark))

        button.bind("<Leave>", lambda _e: button.config(bg=self.primary_color))



    def _set_login_button_state(self, loading: bool):

        if not hasattr(self, "btn_login"):

            return



        self.btn_login.config(

            state="disabled" if loading else "normal",

            text="Conectando..." if loading else "Entrar no painel"

        )



    def _render_login_brand(self, parent: tk.Frame):

        brand_frame = tk.Frame(parent, bg=self.primary_soft, highlightthickness=0, bd=0)

        brand_frame.pack(fill="x")



        tk.Label(

            brand_frame,

            text="GESTÃO INTELIGENTE DE PONTO",

            bg="#DCEFB6",

            fg=self.primary_dark,

            font=("Segoe UI", 8, "bold"),

            padx=12,

            pady=6

        ).pack(anchor="center", pady=(0, 2))



        if self.logo_brand_image:

            tk.Label(

                brand_frame,

                image=self.logo_brand_image,

                bg=self.primary_soft

            ).pack(anchor="center", pady=(0, 2))



        tk.Label(

            brand_frame,

            text="Dixi Auditor",

            bg=self.primary_soft,

            fg=self.text_color,

            font=("Segoe UI", 20, "bold")

        ).pack(anchor="center")



    def _init_login_ui(self):

        for w in self.winfo_children():

            w.destroy()



        try:

            self.state("normal")

        except Exception:

            pass



        self.geometry("740x490")

        self.configure(bg=self.bg_color)



        self.frame = tk.Frame(self, bg=self.bg_color)

        self.frame.pack(expand=True, fill="both")



        # Container Principal Split Card (Pagare Style)

        card_wrapper = tk.Frame(

            self.frame,

            bg=self.surface_color,

            highlightbackground=self.border_color,

            highlightthickness=1,

            bd=0

        )

        card_wrapper.pack(expand=True, fill="both", padx=24, pady=24)



        # Coluna Esquerda: Painel Hero Verde Suave (3.png / 4.png)

        hero_panel = tk.Frame(card_wrapper, bg=self.primary_soft, width=310, bd=0, highlightthickness=0)

        hero_panel.pack(side="left", fill="both", expand=False)

        hero_panel.pack_propagate(False)



        tk.Frame(hero_panel, bg=self.primary_color, width=6).pack(side="left", fill="y")



        hero_content = tk.Frame(hero_panel, bg=self.primary_soft, padx=22, pady=24)

        hero_content.pack(expand=True, fill="both")



        self._render_login_brand(hero_content)



        tk.Label(

            hero_content,

            text="Bem-vindo! 👋",

            bg=self.primary_soft,

            fg=self.text_color,

            font=("Segoe UI", 16, "bold")

        ).pack(anchor="center", pady=(14, 4))



        tk.Label(

            hero_content,

            text="Audite seus pontos, envie justificativas para o RH e acompanhe saldos com facilidade.",

            bg=self.primary_soft,

            fg=self.muted_text_color,

            font=("Segoe UI", 9),

            wraplength=240,

            justify="center"

        ).pack(anchor="center", pady=(0, 14))



        # Bullets de Recursos Centralizados

        features_box = tk.Frame(hero_content, bg=self.primary_soft)

        features_box.pack(anchor="center")



        features = ["✓ Batidas em tempo real", "✓ Auditoria por IA", "✓ Assinatura Autentique"]

        for f in features:

            tk.Label(

                features_box,

                text=f,

                bg=self.primary_soft,

                fg=self.primary_dark,

                font=("Segoe UI", 9, "bold")

            ).pack(anchor="w", pady=1)



        # Coluna Direita: Formulário de Acesso

        login_card = tk.Frame(card_wrapper, bg=self.surface_color, padx=32, pady=24)

        login_card.pack(side="right", fill="both", expand=True)



        tk.Label(

            login_card,

            text="Acessar conta",

            bg=self.surface_color,

            fg=self.text_color,

            font=("Segoe UI", 15, "bold")

        ).pack(anchor="w")





        tk.Label(

            login_card,

            text="Entre com suas credenciais da Dixi para continuar no painel.",

            bg=self.surface_color,

            fg=self.muted_text_color,

            font=("Segoe UI", 9),

            wraplength=260,

            justify="left"

        ).pack(anchor="w", pady=(3, 14))



        self.ent_user = self._build_login_input(login_card, "Usuário / E-mail")

        self.ent_pass = self._build_login_input(login_card, "Senha", show="*")



        last_user = keyring.get_password("DixiPontoApp", "last_user")

        if last_user:

            self.ent_user.insert(0, last_user)

            last_pw = keyring.get_password("DixiPontoApp", last_user)

            if last_pw:

                self.ent_pass.insert(0, last_pw)



        self.lbl_login_status = tk.Label(

            login_card,

            text="",

            bg=self.surface_color,

            fg=self.danger_color,

            font=("Segoe UI", 9)

        )

        self.lbl_login_status.pack(anchor="w", pady=(2, 8))



        self.btn_login = tk.Button(

            login_card,

            text="Entrar no painel",

            command=self._do_login,

            bg=self.primary_color,

            fg="white",

            activebackground=self.primary_dark,

            activeforeground="white",

            font=("Segoe UI", 10, "bold"),

            bd=0,

            relief="flat",

            cursor="hand2",

            padx=14,

            pady=8

        )

        self.btn_login.pack(fill="x", pady=(4, 12))

        self._bind_button_hover(self.btn_login)



        footer = tk.Frame(login_card, bg=self.surface_color)

        footer.pack(fill="x", pady=(6, 0))



        tk.Label(

            footer,

            text="Ambiente protegido",

            bg=self.surface_color,

            fg=self.text_color,

            font=("Segoe UI", 9, "bold")

        ).pack(anchor="center")



        tk.Label(

            footer,

            text="",

            bg=self.surface_color,

            fg=self.muted_text_color,

            font=("Segoe UI", 8),

            wraplength=300,

            justify="center"

        ).pack(anchor="center", pady=(2, 0))



        self.ent_user.focus_set()

        return



        hero_panel = tk.Frame(self.frame, bg=self.primary_soft, bd=0, highlightthickness=0)

        hero_panel.pack(fill="x", pady=(0, 18))



        tk.Frame(hero_panel, bg=self.primary_color, height=4).pack(fill="x")



        hero_body = tk.Frame(hero_panel, bg=self.primary_soft, padx=22, pady=22)

        hero_body.pack(fill="x")

        self._render_login_brand(hero_body)



        login_card = tk.Frame(

            self.frame,

            bg=self.surface_color,

            padx=24,

            pady=24,

            highlightbackground=self.border_color,

            highlightthickness=1,

            bd=0

        )

        login_card.pack(fill="x")



        tk.Label(

            login_card,

            text="Entrar",

            bg=self.surface_color,

            fg=self.text_color,

            font=("Segoe UI", 18, "bold")

        ).pack(anchor="w")



        tk.Label(

            login_card,

            text="Use o mesmo acesso da Dixi para continuar.",

            bg=self.surface_color,

            fg=self.muted_text_color,

            font=("Segoe UI", 10)

        ).pack(anchor="w", pady=(4, 18))



        self.ent_user = self._build_login_input(login_card, "Usuário")

        self.ent_pass = self._build_login_input(login_card, "Senha", show="*")



        last_user = keyring.get_password("DixiPontoApp", "last_user")

        if last_user:

            self.ent_user.insert(0, last_user)

            last_pw = keyring.get_password("DixiPontoApp", last_user)

            if last_pw:

                self.ent_pass.insert(0, last_pw)



        self.btn_login = tk.Button(

            login_card,

            text="Conectar",

            command=self._do_login,

            bg=self.primary_color,

            fg="white",

            activebackground=self.primary_dark,

            activeforeground="white",

            font=("Segoe UI", 11, "bold"),

            bd=0,

            relief="flat",

            cursor="hand2",

            padx=18,

            pady=12

        )

        self.btn_login.pack(fill="x", pady=(6, 12))

        self._bind_button_hover(self.btn_login)



        tk.Label(

            login_card,

            text="Ambiente protegido para consulta de ponto, justificativas e exportações.",

            bg=self.surface_color,

            fg=self.muted_text_color,

            font=("Segoe UI", 9),

            wraplength=340,

            justify="center"

        ).pack(pady=(2, 0))

        return

        self._render_login_brand(self.frame)

        

        lbl_user = tk.Label(self.frame, text="Usuário", bg=self.bg_color, fg="#1E4620", font=("Segoe UI", 11, "bold"))

        lbl_user.pack(anchor="center", pady=(5, 4))

        

        self.ent_user = ttk.Entry(self.frame, width=32, font=("Segoe UI", 10))

        self.ent_user.pack(anchor="center", pady=(0, 14), ipady=4)

        

        lbl_pass = tk.Label(self.frame, text="Senha", bg=self.bg_color, fg="#1E4620", font=("Segoe UI", 11, "bold"))

        lbl_pass.pack(anchor="center", pady=(5, 4))

        

        self.ent_pass = ttk.Entry(self.frame, show="*", width=32, font=("Segoe UI", 10))

        self.ent_pass.pack(anchor="center", pady=(0, 20), ipady=4)



        # RECUPERA O ÚLTIMO LOGIN SALVO

        last_user = keyring.get_password("DixiPontoApp", "last_user")

        if last_user:

            self.ent_user.insert(0, last_user)

            last_pw = keyring.get_password("DixiPontoApp", last_user)

            if last_pw:

                self.ent_pass.insert(0, last_pw)



        self.btn_login = tk.Button(

            self.frame,

            text="Conectar",

            command=self._do_login,

            bg="#5CBD28",

            fg="white",

            activebackground="#4E9E24",

            activeforeground="white",

            font=("Segoe UI", 11, "bold"),

            bd=0,

            relief="flat",

            cursor="hand2",

            width=28,

            pady=8

        )

        self.btn_login.pack(anchor="center", pady=(5, 15))



    def _do_login(self):

        u, p = self.ent_user.get(), self.ent_pass.get()

        self._set_login_button_state(True)

        

        def run():

            if self.service.authenticate(u, p):

                self.after(0, self._init_main_ui)

            else:

                self.after(0, lambda: messagebox.showerror("Erro", "Falha no Login"))

                self.after(0, lambda: self._set_login_button_state(False))

        

        threading.Thread(target=run, daemon=True).start()



    def _init_main_ui(self):

        return self._init_main_ui_modern()



    def _init_main_ui_modern(self):

        for w in self.winfo_children():

            w.destroy()



        self.geometry("1220x760")
        self.minsize(800, 600)

        try:
            self.state("zoomed")
        except Exception:
            pass




        self.configure(bg=self.bg_color)



        self.style.configure(
            "Treeview",
            background="#FFFFFF",
            foreground=self.text_color,
            fieldbackground="#FFFFFF",
            rowheight=30,
            font=("Segoe UI", 10),
            borderColor="#080808",
            bordercolor="#080808",
            lightcolor="#080808",
            darkcolor="#080808",
            relief="solid"
        )

        self.style.configure(
            "Treeview.Heading",
            background=self.primary_soft,
            foreground=self.text_color,
            font=("Segoe UI", 10, "bold"),
            borderColor="#080808",
            bordercolor="#080808",
            lightcolor="#080808",
            darkcolor="#080808",
            relief="solid"
        )




        shell = tk.Frame(self, bg=self.bg_color, padx=18, pady=18)

        shell.pack(fill="both", expand=True)



        header_card = tk.Frame(

            shell,

            bg=self.surface_color,

            padx=22,

            pady=18,

            highlightbackground=self.border_color,

            highlightthickness=1,

            bd=0

        )

        header_card.pack(fill="x", pady=(0, 14))



        header_top = tk.Frame(header_card, bg=self.surface_color)

        header_top.pack(fill="x")



        header_left = tk.Frame(header_top, bg=self.surface_color)

        header_left.pack(side="left", fill="x", expand=True)



        tk.Label(

            header_left,

            text="Dixi Auditor",

            bg=self.surface_color,

            fg=self.text_color,

            font=("Segoe UI", 20, "bold")

        ).pack(anchor="w")



        tk.Label(

            header_left,

            text="Visualize, recalcule e exporte o espelho de ponto em um fluxo mais claro.",

            bg=self.surface_color,

            fg=self.muted_text_color,

            font=("Segoe UI", 10)

        ).pack(anchor="w", pady=(4, 0))



        user_name = self.service.user_name or "Colaborador"

        first_name = user_name.split()[0] if user_name else "Colaborador"

        user_role = self.service.user_cargo or "Acesso autenticado"



        badge = tk.Frame(header_top, bg=self.primary_soft, padx=14, pady=10)

        badge.pack(side="right")



        badge_info = tk.Frame(badge, bg=self.primary_soft)

        badge_info.pack(side="left")



        tk.Label(

            badge_info,

            text=f"Olá, {first_name} 👋",

            bg=self.primary_soft,

            fg=self.text_color,

            font=("Segoe UI", 11, "bold")

        ).pack(anchor="e")







        badge_actions = tk.Frame(badge, bg=self.primary_soft)

        badge_actions.pack(side="right", padx=(12, 0))



        btn_logout = tk.Button(

            badge_actions,

            text="Sair",

            command=self._init_login_ui,

            bg="#FCE8E6",

            fg=self.danger_color,

            activebackground="#F8D7DA",

            activeforeground=self.danger_color,

            font=("Segoe UI", 8, "bold"),

            bd=0,

            relief="flat",

            cursor="hand2",

            padx=10,

            pady=2

        )

        btn_logout.pack(side="top", fill="x", pady=(0, 3))



        btn_config = tk.Button(

            badge_actions,

            text="⚙️ Configurações",

            command=self._abrir_modal_configuracoes,


            bg="#E8F5E9",

            fg=self.primary_dark,

            activebackground="#C8E6C9",

            activeforeground=self.primary_dark,

            font=("Segoe UI", 8, "bold"),

            bd=0,

            relief="flat",

            cursor="hand2",

            padx=8,

            pady=2

        )

        btn_config.pack(side="top", fill="x")



        # Capsule Navbar (Pílula Flutuante de Navegação inspirada em TesteHTML)

        nav_capsule_wrapper = tk.Frame(header_card, bg=self.surface_color)

        nav_capsule_wrapper.pack(fill="x", pady=(14, 0))



        capsule_frame = tk.Frame(
            nav_capsule_wrapper,
            bg=self.primary_soft,
            padx=6,
            pady=6,
            highlightbackground=self.border_color,
            highlightthickness=1,
            bd=0
        )
        capsule_frame.pack(anchor="center")




        self._active_view_key = "ponto"
        self.frame_ponto = tk.Frame(shell, bg=self.bg_color)
        self.frame_ponto.pack(fill="both", expand=True)

        self.frame_justificativa = tk.Frame(shell, bg=self.bg_color)
        self.frame_historico = tk.Frame(shell, bg=self.bg_color)

        self.nav_tabs = {}
        tabs_data = [
            ("ponto", "📊 Espelho de Ponto"),
            ("justificativa", "📝 Justificativas RH"),
            ("historico", "📜 Histórico Autentique")
        ]

        def _select_tab(key):
            self._active_view_key = key
            for t_key, btn in self.nav_tabs.items():
                if t_key == key:
                    btn.config(bg=self.primary_dark, fg="white", font=("Segoe UI", 10, "bold"))
                else:
                    btn.config(bg=self.primary_soft, fg=self.text_color, font=("Segoe UI", 10))

            if key == "ponto":
                self.frame_historico.pack_forget()
                self.frame_justificativa.pack_forget()
                self.frame_ponto.pack(fill="both", expand=True)
            elif key == "justificativa":
                self.frame_ponto.pack_forget()
                self.frame_historico.pack_forget()
                self.frame_justificativa.pack(fill="both", expand=True)
                self._setup_justificativa_ui()
            elif key == "historico":
                self.frame_ponto.pack_forget()
                self.frame_justificativa.pack_forget()
                self.frame_historico.pack(fill="both", expand=True)
                self._carregar_historico_autentique_ui()

        self._select_tab = _select_tab

        for t_key, label_text in tabs_data:
            btn = tk.Button(
                capsule_frame,
                text=label_text,
                command=lambda k=t_key: _select_tab(k),
                bg=self.primary_dark if t_key == "ponto" else self.primary_soft,
                fg="white" if t_key == "ponto" else self.text_color,
                activebackground=self.primary_color,
                activeforeground="white",
                font=("Segoe UI", 10, "bold" if t_key == "ponto" else "normal"),
                bd=0,
                relief="flat",
                cursor="hand2",
                padx=16,
                pady=6
            )
            btn.pack(side="left", padx=3)
            self.nav_tabs[t_key] = btn

        # Configura os widgets da aba Histórico Autentique
        self._setup_historico_ui()

        # Painel de 3 Mini-Cards no topo da aba Ponto (Saldo, Pendências, Período/Preferências)
        metrics_frame = tk.Frame(self.frame_ponto, bg=self.bg_color)
        metrics_frame.pack(fill="x", pady=(0, 14))

        # Card 1: Saldo Acumulado
        card_saldo = tk.Frame(metrics_frame, bg=self.surface_color, padx=14, pady=10, highlightbackground=self.border_color, highlightthickness=1)
        card_saldo.pack(side="left", fill="both", expand=True, padx=(0, 6))

        tk.Label(card_saldo, text="⚖️ Saldo Acumulado", bg=self.surface_color, fg=self.muted_text_color, font=("Segoe UI", 9, "bold")).pack(anchor="w")
        self.lbl_metric_saldo = tk.Label(card_saldo, text="+00:00", bg=self.surface_color, fg=self.primary_dark, font=("Segoe UI", 15, "bold"))
        self.lbl_metric_saldo.pack(anchor="w", pady=(3, 0))

        # Card 2: Pendências / Batidas
        card_pend = tk.Frame(metrics_frame, bg=self.surface_color, padx=14, pady=10, highlightbackground=self.border_color, highlightthickness=1)
        card_pend.pack(side="left", fill="both", expand=True, padx=(3, 6))

        tk.Label(card_pend, text="⚠️ Pendências / Batidas", bg=self.surface_color, fg=self.muted_text_color, font=("Segoe UI", 9, "bold")).pack(anchor="w")
        self.lbl_metric_pendencias = tk.Label(card_pend, text="0 pendências", bg=self.surface_color, fg=self.danger_color, font=("Segoe UI", 15, "bold"))
        self.lbl_metric_pendencias.pack(anchor="w", pady=(3, 0))

        # Card 3: Período de Análise & Preferências
        now = datetime.now()
        current_year = str(now.year)
        current_month = f"{now.month:02d}"
        current_day = f"{now.day:02d}"

        card_period = tk.Frame(metrics_frame, bg=self.surface_color, padx=14, pady=10, highlightbackground=self.border_color, highlightthickness=1)
        card_period.pack(side="left", fill="both", expand=True, padx=(3, 0))

        tk.Label(card_period, text="📅 Período de Análise & Preferências", bg=self.surface_color, fg=self.muted_text_color, font=("Segoe UI", 9, "bold")).pack(anchor="w", pady=(0, 4))

        period_row = tk.Frame(card_period, bg=self.surface_color)
        period_row.pack(anchor="w")

        tk.Label(period_row, text="De:", bg=self.surface_color, fg=self.muted_text_color, font=("Segoe UI", 8)).pack(side="left", padx=(0, 2))
        self.cal_i = DateSelector(period_row, default_day="01", default_month=current_month, default_year=current_year)
        self.cal_i.pack(side="left", padx=(0, 8))

        tk.Label(period_row, text="Até:", bg=self.surface_color, fg=self.muted_text_color, font=("Segoe UI", 8)).pack(side="left", padx=(0, 2))
        self.cal_f = DateSelector(period_row, default_day=current_day, default_month=current_month, default_year=current_year)
        self.cal_f.pack(side="left", padx=(0, 8))

        self.var_ignore_today = tk.BooleanVar(value=True)
        self.chk_ignore_today = ttk.Checkbutton(
            period_row,
            text="Ignorar dia atual",
            variable=self.var_ignore_today,
            command=self._on_toggle_ignore_today
        )
        self.chk_ignore_today.pack(side="left")

        # Tabela do Espelho de Ponto
        table_card = tk.Frame(
            self.frame_ponto,
            bg=self.surface_color,
            padx=16,
            pady=14,
            highlightbackground=self.border_color,
            highlightthickness=1,
            bd=0
        )
        table_card.pack(fill="both", expand=True)

        table_header = tk.Frame(table_card, bg=self.surface_color)
        table_header.pack(fill="x", pady=(0, 10))

        header_left = tk.Frame(table_header, bg=self.surface_color)
        header_left.pack(side="left", anchor="w")

        tk.Label(
            header_left,
            text="Espelho de ponto",
            bg=self.surface_color,
            fg=self.text_color,
            font=("Segoe UI", 12, "bold")
        ).pack(anchor="w")

        tk.Label(
            header_left,
            text="Edite batidas com duplo clique e acompanhe os saldos em tempo real.",
            bg=self.surface_color,
            fg=self.muted_text_color,
            font=("Segoe UI", 9)
        ).pack(anchor="w", pady=(3, 0))

        header_right = tk.Frame(table_header, bg=self.surface_color)
        header_right.pack(side="right", anchor="e")

        self.btn_buscar = ttk.Button(header_right, text="🔍 Visualizar Ponto", command=self._fetch_and_display)
        self.btn_buscar.pack(side="left", padx=(0, 6))

        self.btn_recalc = ttk.Button(header_right, text="🔄 Recalcular Ponto", command=self._recalculate_tree_totals, state="disabled")
        self.btn_recalc.pack(side="left", padx=(0, 6))

        self.btn_export = ttk.Button(header_right, text="📊 Exportar Excel", command=self._export_excel, state="disabled")
        self.btn_export.pack(side="left")




        table_frame = ttk.Frame(table_card)

        table_frame.pack(fill="both", expand=True)



        self.cols = ["Data", "ENTRADA", "SAIDA", "ALMOÇO", "RETORNO", "ENTRADA EXTRA", "SAIDA EXTRA", "TOTAL", "SALDO", "OBS"]

        self.tree = ttk.Treeview(table_frame, columns=self.cols, show="headings", selectmode="extended")

        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)

        hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)

        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky="nsew")

        vsb.grid(row=0, column=1, sticky="ns")

        hsb.grid(row=1, column=0, sticky="ew")

        table_frame.grid_rowconfigure(0, weight=1)

        table_frame.grid_columnconfigure(0, weight=1)

        column_widths = {
            "Data": 115, "ENTRADA": 75, "SAIDA": 75, "ALMOÇO": 75, "RETORNO": 75, "ENTRADA EXTRA": 75, "SAIDA EXTRA": 75,
            "TOTAL": 90, "SALDO": 90, "OBS": 200
        }

        for c in self.cols:
            self.tree.heading(c, text=c, anchor="center")
            self.tree.column(
                c,
                width=column_widths.get(c, 80),
                minwidth=60,
                stretch=True if c in ["Obs", "Data", "ENTRADA", "SAIDA", "ALMOÇO", "RETORNO", "ENTRADA EXTRA", "SAIDA EXTRA"] else False,
                anchor="center"
            )

        def _on_tree_hscroll(event):

            if self.tree.winfo_exists():

                self.tree.xview_scroll(int(-1 * (event.delta / 120)), "units")




        self.tree.bind("<Shift-MouseWheel>", _on_tree_hscroll)

        self.tree.bind("<Double-1>", self.on_double_click)



        self.tree.tag_configure("positive", background="#E6F4EA", foreground="#1E7E34", font=("Segoe UI", 10, "bold"))

        self.tree.tag_configure("negative", background="#FCE8E6", foreground="#C5221F", font=("Segoe UI", 10, "bold"))

        self.tree.tag_configure("missing", background="#FEF7E0", foreground="#B06000", font=("Segoe UI", 10, "bold"))

        self.tree.tag_configure("in_progress", background="#E8F0FE", foreground="#1A73E8", font=("Segoe UI", 10))

        self.tree.tag_configure("normal", background="#FFFFFF", foreground=self.text_color)



        self.lbl_status = tk.Label(

            table_frame,

            text="Dias carregados: 0 | Saldo Acumulado no Período: +00:00",

            bg=self.surface_color,

            fg=self.muted_text_color,

            font=("Segoe UI", 10, "italic"),

            anchor="w"

        )

        self.lbl_status.grid(row=2, column=0, columnspan=2, sticky="ew", pady=(8, 0))



        # Widget Flutuante de Assistente Virtual (Inspirado nas Imagens 9 e 10)

        self._setup_floating_ai_widget()



    def _on_toggle_ignore_today(self):

        if self.processed_data:

            self._refresh_main_table()



    def on_double_click(self, event):

        region = self.tree.identify("region", event.x, event.y)

        if region != "cell":

            return

            

        column = self.tree.identify_column(event.x)

        item = self.tree.identify_row(event.y)

        

        col_idx = int(column[1:]) - 1

        col_name = self.cols[col_idx]

        

        if col_name in ["Sel", "Data", "TOTAL", "SALDO", "OBS"]:

            return

            

        x, y, width, height = self.tree.bbox(item, column)

        

        entry = ttk.Entry(self.tree)

        entry.insert(0, self.tree.set(item, column))

        entry.select_range(0, "end")

        entry.focus_set()

        entry.place(x=x, y=y, width=width, height=height)

        

        def save_edit(event=None):

            new_val = entry.get().strip()

            if new_val:

                try:

                    datetime.strptime(new_val, "%H:%M")

                except ValueError:

                    messagebox.showerror("Formato Inválido", "A hora deve estar no formato de 24h: HH:MM (ex: 08:30 ou 17:45).")

                    entry.destroy()

                    return

            

            self.tree.set(item, column, new_val)

            entry.destroy()



        def cancel_edit(event=None):

            entry.destroy()



        entry.bind("<Return>", save_edit)

        entry.bind("<FocusOut>", save_edit)

        entry.bind("<Escape>", cancel_edit)



    def _fetch_and_display(self):

        try:

            s = self.cal_i.get_date().strftime("%Y%m%d")

            e = self.cal_f.get_date().strftime("%Y%m%d")

        except ValueError as val_ex:

            messagebox.showerror("Erro de Data", str(val_ex))

            return

        

        self.btn_buscar.config(state="disabled")

        self.btn_export.config(state="disabled")

        self.btn_recalc.config(state="disabled")

        if hasattr(self, "btn_ai") and self.btn_ai:

            self.btn_ai.config(state="disabled")

        

        for item in self.tree.get_children():

            self.tree.delete(item)

            

        self.lbl_status.config(text="Obtendo dados do espelho de ponto na Dixi...", foreground=self.primary_dark)

        

        def run():

            try:

                raw = self.service.fetch_history(s, e)

                if not raw:

                    self.after(0, lambda: messagebox.showwarning("Aviso", "Nenhum dado de ponto encontrado para o período selecionado."))

                    return

                

                processed = sorted(

                    [PontoEngine.process_day(d) for d in raw], 

                    key=lambda x: x.data_id

                )

                

                self.processed_data = processed

                self.after(0, lambda: self._populate_table(processed))

                

            except Exception as ex:

                self.after(0, lambda: messagebox.showerror("Erro", f"Erro ao buscar histórico:\n{ex}"))

            finally:

                self.after(0, lambda: self.btn_buscar.config(state="normal"))

        

        threading.Thread(target=run, daemon=True).start()



    def _populate_table(self, data: List[MarcacaoDia]):

        max_horarios = max([len(m.horarios) for m in data]) if data else 0

        max_cols = max(6, max_horarios)

        if max_cols % 2 != 0:

            max_cols += 1



        cols = ["Data"]

        for i in range(1, (max_cols // 2) + 1):

            cols.extend([f"ENTRADA {i}", f"SAIDA {i}"])

        cols.extend(["TOTAL", "SALDO", "OBS"])

        self.cols = cols
        self.tree["columns"] = cols

        for c in cols:
            self.tree.heading(c, text=c, anchor="center")
            if c == "Data":
                self.tree.column(c, width=115, minwidth=80, stretch=True, anchor="center")
            elif c in ["TOTAL", "SALDO"]:
                self.tree.column(c, width=85, minwidth=70, stretch=False, anchor="center")
            elif c == "OBS":
                self.tree.column(c, width=180, minwidth=100, stretch=True, anchor="center")
            else:
                self.tree.column(c, width=70, minwidth=50, stretch=True, anchor="center")

        today_str = datetime.now().strftime("%d/%m/%Y")

        ignore_today = self.var_ignore_today.get()

        for m in data:

            is_today = (m.data_formatada == today_str) and ignore_today

            punches = (m.horarios + [""] * max_cols)[:max_cols]

            total_str = self.exporter.format_time(m.segundos_trabalhados)

            if is_today:

                saldo_str = "00:00"

                obs_str = m.obs if m.obs else "EM ANDAMENTO"

                tag = "in_progress"

            else:

                saldo_str = self.exporter.format_time(m.saldo_segundos, True)

                obs_str = m.obs if m.obs else ("FALTA BATIDA" if m.is_pendencia else "")

                tag = "normal"

                if m.is_pendencia:

                    tag = "missing"

                elif m.saldo_segundos > 0:

                    tag = "positive"

                elif m.saldo_segundos < 0:

                    tag = "negative"

            row_vals = [m.data_formatada] + punches + [total_str, saldo_str, obs_str]

            self.tree.insert("", "end", values=row_vals, tags=(tag,))


            

        if data:

            self.btn_export.config(state="normal")

            self.btn_recalc.config(state="normal")

            if hasattr(self, "btn_ai") and self.btn_ai:

                self.btn_ai.config(state="normal")

            self.btn_justificativa.config(state="normal")

            

            total_trabalhado_seg = sum([

                0 if (m.data_formatada == today_str and ignore_today) else m.segundos_trabalhados

                for m in data

            ])

            total_saldo_seg = sum([

                0 if (m.data_formatada == today_str and ignore_today) else m.saldo_segundos 

                for m in data

            ])

            num_pendencias = sum([

                1 for m in data

                if m.is_pendencia and not (m.data_formatada == today_str and ignore_today)

            ])



            if hasattr(self, "lbl_metric_total"):

                tot_str = self.exporter.format_time(total_trabalhado_seg)

                self.lbl_metric_total.config(text=f"{tot_str}h")

            if hasattr(self, "lbl_metric_saldo"):

                saldo_str = self.exporter.format_time(total_saldo_seg, True)

                color = self.primary_dark if total_saldo_seg >= 0 else self.danger_color

                self.lbl_metric_saldo.config(text=saldo_str, fg=color)

            if hasattr(self, "lbl_metric_pendencias"):

                self.lbl_metric_pendencias.config(

                    text=f"{num_pendencias} pendências" if num_pendencias != 1 else "1 pendência",

                    fg=self.danger_color if num_pendencias > 0 else self.primary_dark

                )



            total_dias = len(data)

            saldo_acumulado = self.exporter.format_time(total_saldo_seg, True)

            

            status_text = f"Dias carregados: {total_dias} | Saldo Acumulado no Período: {saldo_acumulado}"

            if ignore_today and any(m.data_formatada == today_str for m in data):

                status_text += " (Dia atual em andamento desconsiderado no saldo)"

            self.lbl_status.config(text=status_text, foreground=self.text_color)



    def _recalculate_tree_totals(self):

        try:

            total_saldo_seg = 0

            total_dias = 0

            today_str = datetime.now().strftime("%d/%m/%Y")

            ignore_today = self.var_ignore_today.get()

            

            has_sel = ("Sel" in self.cols)

            offset = 1 if has_sel else 0

            num_punch_cols = len(self.cols) - (5 if has_sel else 4)

            new_processed = []

            

            for item in self.tree.get_children():

                values = list(self.tree.item(item, "values"))

                if not values:

                    continue

                

                sel_val = values[0] if has_sel else "[☑]"

                data_formatada = str(values[offset])

                try:

                    dt_obj = datetime.strptime(data_formatada, "%d/%m/%Y")

                    data_id = dt_obj.strftime("%Y%m%d")

                except Exception:

                    continue

                

                punches = [str(values[i]).strip() for i in range(1 + offset, 1 + offset + num_punch_cols) if i < len(values) and str(values[i]).strip()]

                

                day_data = {

                    "data": data_id,

                    "marcacoes": [{"hora": h} for h in punches]

                }

                

                m_dia = PontoEngine.process_day(day_data)

                new_processed.append(m_dia)

                

                is_today = (data_formatada == today_str) and ignore_today

                total_str = self.exporter.format_time(m_dia.segundos_trabalhados)

                

                if is_today:

                    saldo_str = "00:00"

                    obs_str = m_dia.obs if m_dia.obs else "EM ANDAMENTO"

                    tag = "in_progress"

                else:

                    saldo_str = self.exporter.format_time(m_dia.saldo_segundos, True)

                    obs_str = m_dia.obs if m_dia.obs else ("FALTA BATIDA" if m_dia.is_pendencia else "")

                    tag = "normal"

                    if m_dia.is_pendencia:

                        tag = "missing"

                    elif m_dia.saldo_segundos > 0:

                        tag = "positive"

                    elif m_dia.saldo_segundos < 0:

                        tag = "negative"

                    

                padded_punches = (punches + [""] * num_punch_cols)[:num_punch_cols]

                new_values = ([sel_val] if has_sel else []) + [data_formatada] + padded_punches + [total_str, saldo_str, obs_str]

                self.tree.item(item, values=new_values, tags=(tag,))

                

                if len(punches) > 0 and not is_today:

                    total_saldo_seg += m_dia.saldo_segundos

                    total_dias += 1

                    

            self.processed_data = new_processed

            

            total_trabalhado_seg = sum([

                0 if (m.data_formatada == today_str and ignore_today) else m.segundos_trabalhados

                for m in new_processed

            ])

            num_pendencias = sum([

                1 for m in new_processed

                if m.is_pendencia and not (m.data_formatada == today_str and ignore_today)

            ])



            if hasattr(self, "lbl_metric_total"):

                tot_str = self.exporter.format_time(total_trabalhado_seg)

                self.lbl_metric_total.config(text=f"{tot_str}h")

            if hasattr(self, "lbl_metric_saldo"):

                saldo_str = self.exporter.format_time(total_saldo_seg, True)

                color = self.primary_dark if total_saldo_seg >= 0 else self.danger_color

                self.lbl_metric_saldo.config(text=saldo_str, fg=color)

            if hasattr(self, "lbl_metric_pendencias"):

                self.lbl_metric_pendencias.config(

                    text=f"{num_pendencias} pendências" if num_pendencias != 1 else "1 pendência",

                    fg=self.danger_color if num_pendencias > 0 else self.primary_dark

                )

            

            saldo_acumulado = self.exporter.format_time(total_saldo_seg, True)

            status_text = f"🔄 Ponto Recalculado! Dias: {total_dias} | Saldo Acumulado: {saldo_acumulado}"

            if ignore_today and any(m.data_formatada == today_str for m in new_processed):

                status_text += " (Dia atual desconsiderado)"

            self.lbl_status.config(text=status_text, foreground=self.primary_dark)

            messagebox.showinfo("Recálculo Concluído", f"Ponto recalculado com sucesso para {total_dias} dias!\n\nSaldo Acumulado no Período: {saldo_acumulado}", parent=self)

        except Exception as ex:

            logging.error(f"Erro ao recalcular ponto: {ex}")

            messagebox.showerror("Erro de Recálculo", f"Falha ao recalcular o ponto:\n{ex}", parent=self)



    def _export_excel(self):

        if not self.processed_data:

            messagebox.showwarning("Aviso", "Nenhum dado carregado para exportação.")

            return

            

        try:

            s = self.cal_i.get_date().strftime("%Y%m%d")

            e = self.cal_f.get_date().strftime("%Y%m%d")

        except ValueError as val_ex:

            messagebox.showerror("Erro de Data", str(val_ex))

            return

        

        path = filedialog.asksaveasfilename(

            defaultextension=".xlsx",

            filetypes=[("Planilha Excel", "*.xlsx")],

            initialfile=f"Ponto_{s}_a_{e}.xlsx",

            title="Salvar Relatório de Ponto"

        )

        if path:

            try:

                self.exporter.generate(self.processed_data, path, ignore_today=self.var_ignore_today.get())

                messagebox.showinfo("Sucesso", "Planilha Excel gerada com sucesso!")

                os.startfile(path)

            except Exception as ex_save:

                messagebox.showerror("Erro ao Salvar", f"Não foi possível salvar a planilha:\n{ex_save}")



    def _config_ai_key(self):

        top = tk.Toplevel(self)

        top.title("Configurar Chave de IA (OpenRouter)")

        top.geometry("480x240")

        top.configure(bg=self.bg_color)

        top.transient(self)

        top.resizable(False, False)



        tk.Label(

            top, 

            text="🔑 Configuração da Chave da IA", 

            bg=self.bg_color, 

            fg=self.text_color, 

            font=("Segoe UI", 12, "bold")

        ).pack(pady=(15, 5))



        tk.Label(

            top, 

            text="Cole abaixo sua chave do OpenRouter (ex: sk-or-v1-...) para habilitar as análises avançadas de IA:", 

            bg=self.bg_color, 

            fg="#2E4E2E", 

            font=("Segoe UI", 9),

            wraplength=440,

            justify="center"

        ).pack(padx=15, pady=(0, 10))



        curr_key = IAAnalistaPonto.get_api_key()



        ent_key = ttk.Entry(top, width=50, show="*")

        if curr_key:

            ent_key.insert(0, curr_key)

        ent_key.pack(pady=(0, 15), ipady=3)



        lbl_info = tk.Label(top, text="", bg=self.bg_color, fg=self.primary_dark, font=("Segoe UI", 9, "italic"))

        lbl_info.pack(pady=(0, 5))



        def save():

            k = ent_key.get().strip()

            if k:

                keyring.set_password("DixiPontoApp", "openrouter_token", k)

                os.environ["ANTHROPIC_AUTH_TOKEN"] = k

                lbl_info.config(text="✅ Chave salva com segurança no Windows Keyring!", fg=self.primary_dark)

            else:

                try:

                    keyring.delete_password("DixiPontoApp", "openrouter_token")

                except Exception:

                    pass

                os.environ.pop("ANTHROPIC_AUTH_TOKEN", None)

                lbl_info.config(text="ℹ️ Chave removida.", fg=self.danger_color)

            self.after(1500, top.destroy)



        btn_save = tk.Button(

            top,

            text="Salvar Chave",

            command=save,

            bg="#5CBD28",

            fg="white",

            font=("Segoe UI", 10, "bold"),

            bd=0,

            relief="flat",

            cursor="hand2",

            padx=15,

            pady=5

        )

        btn_save.pack()



    def _apply_ai_adjustments(self, ajustes: List[Dict]) -> int:

        count = 0

        for aj in ajustes:

            dt_target = str(aj.get("data", "")).strip()

            if not dt_target:

                continue

            

            for m in self.processed_data:

                if m.data_formatada == dt_target or m.data_formatada.startswith(dt_target):

                    if aj.get("abono"):

                        m.is_pendencia = False

                        m.obs = aj.get("obs", "Abonado via IA")

                        count += 1

                    elif aj.get("horarios"):

                        novos_horarios = sorted(aj["horarios"])

                        obs = aj.get("obs", "Ajustado via IA")

                        m_novo = PontoEngine.process_horarios(novos_horarios, m.data_id, m.data_formatada, obs=obs)

                        m.segundos_trabalhados = m_novo.segundos_trabalhados

                        m.saldo_segundos = m_novo.saldo_segundos

                        m.is_pendencia = m_novo.is_pendencia

                        m.horarios = m_novo.horarios

                        m.obs = obs

                        count += 1



        if count > 0:

            self._refresh_main_table()

        return count



    def _setup_floating_ai_widget(self):

        self.ai_widget_frame = tk.Frame(self, bg=self.bg_color)

        self.ai_widget_frame.place(relx=0.98, rely=0.96, anchor="se")



        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

        img_circle_path = os.path.join(base_dir, "assets", "images", "ai_avatar_circle.png")

        img_header_path = os.path.join(base_dir, "assets", "images", "ai_avatar_header.png")



        self._ai_avatar_img = None

        self._ai_header_img = None



        if os.path.isfile(img_circle_path):

            try:

                self._ai_avatar_img = ImageTk.PhotoImage(Image.open(img_circle_path))

            except Exception:

                pass

        if os.path.isfile(img_header_path):

            try:

                self._ai_header_img = ImageTk.PhotoImage(Image.open(img_header_path))

            except Exception:

                pass



        # 1. Balão de fala + Botão Avatar Circular (Fiel à Imagem 9)

        self.ai_bubble_container = tk.Frame(self.ai_widget_frame, bg=self.bg_color)

        self.ai_bubble_container.pack(side="right")



        speech_bubble = tk.Frame(

            self.ai_bubble_container,

            bg="#FFFFFF",

            padx=14,

            pady=10,

            highlightbackground="#CBD5E1",

            highlightthickness=1,

            bd=0,

            cursor="hand2"

        )

        speech_bubble.pack(side="left", padx=(0, 10))



        lbl_bubble_txt = tk.Label(

            speech_bubble,

            text="Precisa de ajuda com a auditoria\ndo seu ponto?",

            bg="#FFFFFF",

            fg="#1E293B",

            font=("Segoe UI", 9, "bold"),

            justify="left"

        )

        lbl_bubble_txt.pack()



        if self._ai_avatar_img:

            btn_avatar = tk.Label(

                self.ai_bubble_container,

                image=self._ai_avatar_img,

                bg=self.bg_color,

                cursor="hand2"

            )

        else:

            btn_avatar = tk.Label(

                self.ai_bubble_container,

                text="🤖",

                bg=self.primary_dark,

                fg="white",

                font=("Segoe UI", 16),

                padx=12,

                pady=12,

                cursor="hand2"

            )

        btn_avatar.pack(side="left")



        # 2. Popup do Chat de Assistente Virtual (Fiel à Imagem 10) - Tamanho Compacto (380x500px)

        self.ai_chat_popup = tk.Frame(

            self.ai_widget_frame,

            bg="#FFFFFF",

            width=380,

            height=550,

            highlightbackground="#6acc32",

            highlightthickness=2,

            bd=0

        )

        self.ai_chat_popup.pack_propagate(False)



        # 1. Header do Chat (Verde #2E7D32)

        chat_header = tk.Frame(self.ai_chat_popup, bg="#2E7D32", padx=12, pady=10)

        chat_header.pack(fill="x", side="top")



        if self._ai_header_img:

            lbl_hdr_avatar = tk.Label(chat_header, image=self._ai_header_img, bg="#2E7D32")

            lbl_hdr_avatar.pack(side="left", padx=(0, 8))

        else:

            lbl_hdr_avatar = tk.Label(chat_header, text="🤖", bg="#2E7D32", fg="white", font=("Segoe UI", 14))

            lbl_hdr_avatar.pack(side="left", padx=(0, 8))



        title_box = tk.Frame(chat_header, bg="#2E7D32")

        title_box.pack(side="left", fill="both", expand=True)



        tk.Label(

            title_box,

            text="Assistente Virtual - Dixi Auditor",

            bg="#2E7D32",

            fg="#FFFFFF",

            font=("Segoe UI", 10, "bold")

        ).pack(anchor="w")



        tk.Label(

            title_box,

            text="Online agora 🟢",

            bg="#2E7D32",

            fg="#A7F3D0",

            font=("Segoe UI", 8)

        ).pack(anchor="w")



        btn_close_chat = tk.Button(

            chat_header,

            text="✕",

            command=self._close_ai_chat,

            bg="#2E7D32",

            fg="#FFFFFF",

            activebackground="#1B5E20",

            activeforeground="#FFFFFF",

            font=("Segoe UI", 11, "bold"),

            bd=0,

            relief="flat",

            cursor="hand2",

            padx=6

        )

        btn_close_chat.pack(side="right")



        # 2. Footer com campo de texto e envio FIXO no rodapé

        chat_footer = tk.Frame(self.ai_chat_popup, bg="#FFFFFF", padx=10, pady=8, highlightbackground="#E2E8F0", highlightthickness=1, bd=0)

        chat_footer.pack(fill="x", side="bottom")



        self.ent_chat_msg = ttk.Entry(chat_footer, font=("Segoe UI", 9))

        self.ent_chat_msg.pack(side="left", fill="x", expand=True, padx=(0, 6))

        self.ent_chat_msg.bind("<Return>", lambda e: self._send_ai_chat_msg())



        btn_send = tk.Button(

            chat_footer,

            text="➔",

            command=self._send_ai_chat_msg,

            bg="#2E7D32",

            fg="#FFFFFF",

            activebackground="#15803D",

            activeforeground="#FFFFFF",

            font=("Segoe UI", 9, "bold"),

            bd=0,

            relief="flat",

            cursor="hand2",

            padx=10,

            pady=2

        )

        btn_send.pack(side="right")



        # 3. Banners rápidos de Ação acima do footer (Somente Auditar Ponto)

        quick_actions_frame = tk.Frame(self.ai_chat_popup, bg="#F8FAFC", padx=10, pady=4)

        quick_actions_frame.pack(fill="x", side="bottom")



        btn_audit_quick = tk.Button(

            quick_actions_frame,

            text="🤖 Auditar Ponto do Período",

            command=self._run_quick_ai_audit,

            bg="#2E7D32",

            fg="#FFFFFF",

            activebackground="#15803D",

            activeforeground="#FFFFFF",

            font=("Segoe UI", 8, "bold"),

            bd=0,

            relief="flat",

            cursor="hand2",

            padx=8,

            pady=4

        )

        btn_audit_quick.pack(side="left", padx=(0, 6))



        btn_sugestao_quick = tk.Button(

            quick_actions_frame,

            text="✨ Aplicar Sugestão",

            command=self._apply_ai_suggestion_button,

            bg="#0284C7",

            fg="#FFFFFF",

            activebackground="#0369A1",

            activeforeground="#FFFFFF",

            font=("Segoe UI", 8, "bold"),

            bd=0,

            relief="flat",

            cursor="hand2",

            padx=8,

            pady=4

        )

        btn_sugestao_quick.pack(side="left")



        # 4. Corpo de Mensagens do Chat (Preenche todo o meio)

        chat_body = tk.Frame(self.ai_chat_popup, bg="#F8FAFC", padx=10, pady=10)

        chat_body.pack(fill="both", expand=True, side="top")



        self.txt_chat_history = tk.Text(

            chat_body,

            font=("Segoe UI", 9),

            wrap="word",

            bg="#F8FAFC",

            fg="#0F172A",

            relief="flat",

            bd=0

        )

        vsb_chat = ttk.Scrollbar(chat_body, orient="vertical", command=self.txt_chat_history.yview)

        self.txt_chat_history.configure(yscrollcommand=vsb_chat.set)



        self.txt_chat_history.pack(side="left", fill="both", expand=True)

        vsb_chat.pack(side="right", fill="y")



        # Binds para abrir o chat ao clicar no avatar ou balão

        speech_bubble.bind("<Button-1>", lambda e: self._open_ai_chat())

        lbl_bubble_txt.bind("<Button-1>", lambda e: self._open_ai_chat())

        btn_avatar.bind("<Button-1>", lambda e: self._open_ai_chat())



        # Mensagem inicial do assistente (Fiel à imagem 10)

        self.txt_chat_history.insert("1.0", "🤖 Assistente Dixi: Olá, tudo bem?\n\nSou a assistente virtual do Dixi Auditor. Como posso te ajudar hoje?\n\n• Clique no botão verde '🤖 Auditar Ponto' para auditar automaticamente faltas e horas extras.\n• Ou digite sua dúvida no campo abaixo!\n\n" + "─"*35 + "\n\n")

        self.txt_chat_history.config(state="disabled")



    def _open_ai_chat(self):

        self.ai_bubble_container.pack_forget()

        self.ai_chat_popup.pack(side="right")



    def _close_ai_chat(self):

        self.ai_chat_popup.pack_forget()

        self.ai_bubble_container.pack(side="right")



    def _apply_ai_suggestion_button(self):

        if not self.processed_data:

            self.txt_chat_history.config(state="normal")

            self.txt_chat_history.insert("end", "\n🤖 Assistente Dixi: Primeiro clique em 'Visualizar Ponto' na tela principal para carregar os registros.\n\n")

            self.txt_chat_history.see("end")

            self.txt_chat_history.config(state="disabled")

            return



        sugestoes_ajustes, texto_sug = IAAnalistaPonto.encontrar_sugestoes_duplicadas(self.processed_data)

        if sugestoes_ajustes:

            cnt = self._apply_ai_adjustments(sugestoes_ajustes)

            self.txt_chat_history.config(state="normal")

            self.txt_chat_history.insert("end", f"\n✨ **SUGESTÃO DE CORREÇÃO APLICADA!**\nForam corrigidos e unificados {cnt} dia(s) com batidas duplicadas no espelho de ponto.\n\n" + "─"*35 + "\n\n")

            self.txt_chat_history.see("end")

            self.txt_chat_history.config(state="disabled")

        else:

            self.txt_chat_history.config(state="normal")

            self.txt_chat_history.insert("end", "\n🤖 Assistente Dixi: Nenhuma batida duplicada foi encontrada no período atual. O ponto está regular!\n\n" + "─"*35 + "\n\n")

            self.txt_chat_history.see("end")

            self.txt_chat_history.config(state="disabled")



    def _run_quick_ai_audit(self):

        if not self.processed_data:

            self.txt_chat_history.config(state="normal")

            self.txt_chat_history.insert("end", "\n🤖 Assistente Dixi: Para auditar o ponto, primeiro clique em 'Visualizar Ponto' na tela principal para carregar os registros.\n\n")

            self.txt_chat_history.see("end")

            self.txt_chat_history.config(state="disabled")

            return



        self.txt_chat_history.config(state="normal")

        self.txt_chat_history.insert("end", "\n⏳ Processando análise do ponto com a IA, aguarde...\n")

        self.txt_chat_history.see("end")

        self.txt_chat_history.config(state="disabled")



        def _worker():

            res_text, ajustes, _ = IAAnalistaPonto.analisar_ponto(

                self.processed_data,

                ignore_today=self.var_ignore_today.get()

            )

            def _update_ui():

                if ajustes:

                    cnt = self._apply_ai_adjustments(ajustes)

                    adj_msg = f"\n\n✨ {cnt} {'ajuste aplicado' if cnt == 1 else 'ajustes aplicados'} automaticamente na tabela de ponto!"

                else:

                    adj_msg = ""



                self.txt_chat_history.config(state="normal")

                self.txt_chat_history.insert("end", f"\n🤖 AUDITORIA AUTOMÁTICA:\n{res_text}{adj_msg}\n\n" + "─"*35 + "\n\n")

                self.txt_chat_history.see("end")

                self.txt_chat_history.config(state="disabled")

            self.after(0, _update_ui)



        threading.Thread(target=_worker, daemon=True).start()



    def _send_ai_chat_msg(self):

        msg = self.ent_chat_msg.get().strip()

        if not msg:

            return

        self.ent_chat_msg.delete(0, "end")



        self.txt_chat_history.config(state="normal")

        self.txt_chat_history.insert("end", f"\n👤 Você: {msg}\n")

        self.txt_chat_history.see("end")

        self.txt_chat_history.config(state="disabled")



        def _worker():

            if self.processed_data:

                res_text, ajustes, auto_enviar = IAAnalistaPonto.analisar_ponto(

                    self.processed_data,

                    instrucoes_usuario=msg,

                    ignore_today=self.var_ignore_today.get()

                )

                res_str = res_text

            else:

                ajustes = []

                auto_enviar = False

                res_str = "Para responder com base nos dados do ponto, por favor clique em 'Visualizar Ponto' na tela principal para carregar seu espelho de ponto."



            def _update_ui():

                if ajustes:

                    cnt = self._apply_ai_adjustments(ajustes)

                    adj_msg = f"\n\n✨ {cnt} {'ajuste aplicado' if cnt == 1 else 'ajustes aplicados'} na tabela do espelho de ponto!"

                else:

                    adj_msg = ""



                self.txt_chat_history.config(state="normal")

                self.txt_chat_history.insert("end", f"\n🤖 Assistente Dixi: {res_str}{adj_msg}\n\n" + "─"*35 + "\n\n")

                self.txt_chat_history.see("end")

                self.txt_chat_history.config(state="disabled")



                if auto_enviar and hasattr(self, "btn_justificativa") and self.btn_justificativa["state"] != "disabled":

                    self._abrir_modal_justificativa()



            self.after(0, _update_ui)



        threading.Thread(target=_worker, daemon=True).start()



    def _refresh_main_table(self):

        for item in self.tree.get_children():

            self.tree.delete(item)

        self._populate_table(self.processed_data)



    def _show_ai_analysis(self):
        top = tk.Toplevel(self)
        top.title("Assistente Virtual de IA - Dixi Auditor")
        top.minsize(680, 500)
        top.configure(bg=self.bg_color)
        top.transient(self)

        # Centraliza a janela 800x600 na tela
        w, h = 800, 600
        ws = top.winfo_screenwidth()
        hs = top.winfo_screenheight()
        x = max(0, (ws // 2) - (w // 2))
        y = max(0, (hs // 2) - (h // 2))
        top.geometry(f"{w}x{h}+{x}+{y}")



        hdr_frame = ttk.Frame(top, padding=(15, 10))

        hdr_frame.pack(fill="x")



        tk.Label(

            hdr_frame, 

            text="🤖 Análise Inteligente de Ponto", 

            bg=self.bg_color, 

            fg=self.text_color, 

            font=("Segoe UI", 14, "bold")

        ).pack(side="left")



        btn_cfg = tk.Button(

            hdr_frame,

            text="🔑 Configurar Chave",

            command=self._config_ai_key,

            bg=self.primary_soft,

            fg=self.text_color,

            font=("Segoe UI", 9, "bold"),

            bd=1,

            relief="solid",

            cursor="hand2",

            padx=8,

            pady=3

        )

        btn_cfg.pack(side="right")



        txt_frame = ttk.Frame(top, padding=(15, 5, 15, 5))

        txt_frame.pack(fill="both", expand=True)



        txt = tk.Text(txt_frame, font=("Segoe UI", 10), wrap="word", bg="#FFFFFF", fg="#1E3014", relief="solid", bd=1)

        vsb = ttk.Scrollbar(txt_frame, orient="vertical", command=txt.yview)

        txt.configure(yscrollcommand=vsb.set)



        txt.pack(side="left", fill="both", expand=True)

        vsb.pack(side="right", fill="y")



        # Container inferior para solicitar edições e recalcular

        edit_frame = tk.LabelFrame(top, text=" ✏️ Editar Pontos de Atenção & Recalcular ", bg=self.bg_color, fg=self.primary_color, font=("Segoe UI", 10, "bold"), padx=10, pady=8)

        edit_frame.pack(fill="x", padx=15, pady=(5, 15))



        lbl_instruct = tk.Label(

            edit_frame,

            text="Descreva as alterações desejadas nos pontos de atenção (ex: 'No dia 15/07 considere saída às 18:00 e abone o dia 10/07'):",

            bg=self.bg_color,

            fg=self.text_color,

            font=("Segoe UI", 9)

        )

        lbl_instruct.pack(anchor="w", pady=(0, 4))



        input_container = ttk.Frame(edit_frame)

        input_container.pack(fill="x")



        ent_instrucao = ttk.Entry(input_container, font=("Segoe UI", 10))

        ent_instrucao.pack(side="left", fill="x", expand=True, padx=(0, 8))



        btn_recalcular = tk.Button(

            input_container,

            text="🔄 Recalcular com IA",

            bg=self.primary_color,

            fg="#FFFFFF",

            font=("Segoe UI", 9, "bold"),

            bd=0,

            relief="flat",

            cursor="hand2",

            padx=12,

            pady=5

        )

        btn_recalcular.pack(side="right")



        lbl_status = tk.Label(edit_frame, text="", bg=self.bg_color, fg="#555555", font=("Segoe UI", 8, "italic"))

        lbl_status.pack(anchor="w", pady=(4, 0))



        def execute_analysis(custom_prompt=None):

            if not self.processed_data:

                txt.config(state="normal")

                txt.delete("1.0", "end")

                txt.insert("1.0", "🤖 Olá! Sou o seu Assistente Virtual de Ponto Dixi Auditor.\n\nNenhum espelho de ponto está selecionado no momento.\n\n• Clique no botão 'Visualizar Ponto' na tela principal para carregar o período desejado.\n• Em seguida, clique novamente em mim para auditar automaticamente faltas, horas extras, banco de horas e inconsistências!\n\n💡 Dica: Você também pode cadastrar ou alterar sua Chave de API no botão '🔑 Configurar Chave' no canto superior direito.")

                txt.config(state="disabled")

                return



            btn_recalcular.config(state="disabled")

            lbl_status.config(text="⏳ Processando análise e recalculando com a IA...")

            if custom_prompt:

                txt.config(state="normal")

                txt.insert("end", f"\n\n{'='*55}\n✏️ INSTRUÇÃO DE AJUSTE ENVIADA:\n\"{custom_prompt}\"\n{'='*55}\n\nRecalculando com IA...\n")

                txt.see("end")

                txt.config(state="disabled")

            else:

                txt.config(state="normal")

                txt.delete("1.0", "end")

                txt.insert("1.0", "Analisando dados do ponto com IA, aguarde alguns segundos...\n")

                txt.config(state="disabled")



            def run_thread():

                res_text, ajustes, auto_enviar = IAAnalistaPonto.analisar_ponto(

                    self.processed_data, 

                    instrucoes_usuario=custom_prompt, 

                    ignore_today=self.var_ignore_today.get()

                )

                def update_ui():

                    txt.config(state="normal")

                    if custom_prompt:

                        txt.insert("end", f"\n🤖 NOVA ANÁLISE RECALCULADA:\n{res_text}\n")

                    else:

                        txt.delete("1.0", "end")

                        txt.insert("1.0", res_text)

                    txt.see("end")

                    txt.config(state="disabled")

                    btn_recalcular.config(state="normal")

                    

                    if ajustes:

                        modificados = self._apply_ai_adjustments(ajustes)

                        if modificados > 0:

                            lbl_status.config(text=f"✅ Análise recalculada e {modificados} dia(s) atualizado(s) na tabela principal do aplicativo!")

                        else:

                            lbl_status.config(text="✅ Análise recalculada com sucesso.")

                    else:

                        lbl_status.config(text="✅ Análise recalculada com sucesso.")

                        

                    ent_instrucao.delete(0, "end")



                    if auto_enviar:

                        lbl_status.config(text="🚀 IA abrindo formulário de Justificativa para o RH...")

                        self.after(600, self._abrir_modal_justificativa)

                self.after(0, update_ui)



            threading.Thread(target=run_thread, daemon=True).start()



        def on_recalculou():

            instrucao = ent_instrucao.get().strip()

            if not instrucao:

                messagebox.showwarning("Aviso", "Digite a instrução ou ajuste que deseja que a IA aplique no recálculo.", parent=top)

                return

            execute_analysis(instrucao)



        btn_recalcular.config(command=on_recalculou)

        ent_instrucao.bind("<Return>", lambda e: on_recalculou())



        # Executa análise inicial sem instrução prévia

        execute_analysis(None)



    def _abrir_modal_justificativa(self):
        if hasattr(self, "_select_tab"):
            self._select_tab("justificativa")





    def _abrir_modal_configuracoes(self):
        top = tk.Toplevel(self)
        top.title("⚙️ Configurações do Sistema")
        top.geometry("720x620")
        top.minsize(650, 520)
        top.transient(self)
        top.grab_set()

        def safe_keyring_get(key, default=""):
            try:
                return keyring.get_password("DixiPontoApp", key) or default
            except Exception:
                return default

        notebook = ttk.Notebook(top)
        notebook.pack(fill="both", expand=True, padx=12, pady=12)

        # Aba 1: 👤 Dados do Colaborador & Aprovadores
        tab_colab = ttk.Frame(notebook, padding=15)
        notebook.add(tab_colab, text="👤 Dados do Colaborador")

        sec_c = ttk.LabelFrame(tab_colab, text=" Dados Pessoais do Colaborador ", padding=12)
        sec_c.pack(fill="x", pady=(0, 10))
        sec_c.columnconfigure(1, weight=1)

        dixi_srv = getattr(self, "service", None) or getattr(self, "dixi", None)
        dixi_email_def = getattr(dixi_srv, "user_email", "") if dixi_srv else ""
        dixi_name_def = getattr(dixi_srv, "user_name", "") if dixi_srv else ""

        ttk.Label(sec_c, text="Nome Completo:").grid(row=0, column=0, sticky="w", pady=4)
        ent_colab_nome = ttk.Entry(sec_c)
        ent_colab_nome.insert(0, safe_keyring_get("colaborador_nome", dixi_name_def or "Colaborador"))
        ent_colab_nome.grid(row=0, column=1, sticky="ew", pady=4, padx=(8, 0))

        ttk.Label(sec_c, text="Cargo / Função:").grid(row=1, column=0, sticky="w", pady=4)
        ent_colab_cargo = ttk.Entry(sec_c)
        ent_colab_cargo.insert(0, safe_keyring_get("colaborador_cargo", ""))
        ent_colab_cargo.grid(row=1, column=1, sticky="ew", pady=4, padx=(8, 0))

        ttk.Label(sec_c, text="E-mail do Colaborador:").grid(row=2, column=0, sticky="w", pady=4)
        ent_colab_email = ttk.Entry(sec_c)
        ent_colab_email.insert(0, safe_keyring_get("colaborador_email", dixi_email_def))
        ent_colab_email.grid(row=2, column=1, sticky="ew", pady=4, padx=(8, 0))

        sec_ap = ttk.LabelFrame(tab_colab, text=" Aprovadores e Destinatários (Gestão / RH) ", padding=12)
        sec_ap.pack(fill="x", pady=(0, 10))
        sec_ap.columnconfigure(1, weight=1)

        ttk.Label(sec_ap, text="Nome do Gestor:").grid(row=0, column=0, sticky="w", pady=4)
        ent_gestor_nome = ttk.Entry(sec_ap)
        ent_gestor_nome.insert(0, safe_keyring_get("gestor_nome", "Gestor Imediato"))
        ent_gestor_nome.grid(row=0, column=1, sticky="ew", pady=4, padx=(8, 0))

        ttk.Label(sec_ap, text="E-mail do Gestor:").grid(row=1, column=0, sticky="w", pady=4)
        ent_gestor_email = ttk.Entry(sec_ap)
        ent_gestor_email.insert(0, safe_keyring_get("gestor_email", ""))
        ent_gestor_email.grid(row=1, column=1, sticky="ew", pady=4, padx=(8, 0))

        ttk.Label(sec_ap, text="Nome do RH:").grid(row=2, column=0, sticky="w", pady=4)
        ent_rh_nome = ttk.Entry(sec_ap)
        ent_rh_nome.insert(0, safe_keyring_get("rh_nome", "Recursos Humanos"))
        ent_rh_nome.grid(row=2, column=1, sticky="ew", pady=4, padx=(8, 0))

        ttk.Label(sec_ap, text="E-mail do RH:").grid(row=3, column=0, sticky="w", pady=4)
        ent_rh_email = ttk.Entry(sec_ap)
        ent_rh_email.insert(0, safe_keyring_get("rh_email", ""))
        ent_rh_email.grid(row=3, column=1, sticky="ew", pady=4, padx=(8, 0))

        lbl_colab_msg = ttk.Label(tab_colab, text="", font=("Segoe UI", 9, "bold"))
        lbl_colab_msg.pack(anchor="w", pady=(0, 8))

        def salvar_dados_colab():
            try:
                keyring.set_password("DixiPontoApp", "colaborador_nome", ent_colab_nome.get().strip())
                keyring.set_password("DixiPontoApp", "colaborador_cargo", ent_colab_cargo.get().strip())
                keyring.set_password("DixiPontoApp", "colaborador_email", ent_colab_email.get().strip())
                keyring.set_password("DixiPontoApp", "gestor_nome", ent_gestor_nome.get().strip())
                keyring.set_password("DixiPontoApp", "gestor_email", ent_gestor_email.get().strip())
                keyring.set_password("DixiPontoApp", "rh_nome", ent_rh_nome.get().strip())
                keyring.set_password("DixiPontoApp", "rh_email", ent_rh_email.get().strip())
                lbl_colab_msg.config(text="✅ Dados do Colaborador salvos com sucesso!", foreground="#16a34a")
            except Exception as ex_s:
                lbl_colab_msg.config(text=f"❌ Erro ao salvar: {ex_s}", foreground="#dc2626")

        btn_save_colab = ttk.Button(tab_colab, text="💾 Salvar Dados do Colaborador", command=salvar_dados_colab)
        btn_save_colab.pack(anchor="e")

        # Aba 2: 🔑 Chaves API & Autentique
        tab_api = ttk.Frame(notebook, padding=15)
        notebook.add(tab_api, text="🔑 Chaves API & Autentique")

        sec_keys = ttk.LabelFrame(tab_api, text=" Credenciais de Integração ", padding=12)
        sec_keys.pack(fill="x", pady=(0, 10))
        sec_keys.columnconfigure(1, weight=1)

        ttk.Label(sec_keys, text="Token Autentique:").grid(row=0, column=0, sticky="w", pady=4)
        ent_autentique_tok = ttk.Entry(sec_keys, show="*")
        ent_autentique_tok.insert(0, safe_keyring_get("autentique_token", ""))
        ent_autentique_tok.grid(row=0, column=1, sticky="ew", pady=4, padx=(8, 0))

        sec_pos = ttk.LabelFrame(tab_api, text=" Posição das Assinaturas no PDF ", padding=12)
        sec_pos.pack(fill="x", pady=(0, 10))
        sec_pos.columnconfigure(1, weight=1)

        saved_preset = safe_keyring_get("autentique_pos_preset", "Sobre a Linha Verde (Y: 68%)")
        ttk.Label(sec_pos, text="Preset de Posição:").grid(row=0, column=0, sticky="w", pady=4)
        combo_preset = ttk.Combobox(sec_pos, values=["Sobre a Linha Verde (Y: 68%)", "Rodapé (Y: 85%)", "Personalizado"], state="readonly")
        combo_preset.set(saved_preset)
        combo_preset.grid(row=0, column=1, sticky="ew", pady=4, padx=(8, 0))

        lbl_api_msg = ttk.Label(tab_api, text="", font=("Segoe UI", 9, "bold"))
        lbl_api_msg.pack(anchor="w", pady=(0, 8))

        def salvar_chaves_api():
            try:
                tok = ent_autentique_tok.get().strip()
                if tok:
                    keyring.set_password("DixiPontoApp", "autentique_token", tok)
                keyring.set_password("DixiPontoApp", "autentique_pos_preset", combo_preset.get())
                lbl_api_msg.config(text="✅ Configurações de API salvas com sucesso!", foreground="#16a34a")
            except Exception as ex_api:
                lbl_api_msg.config(text=f"❌ Erro ao salvar: {ex_api}", foreground="#dc2626")

        btn_save_api = ttk.Button(tab_api, text="💾 Salvar Configurações API", command=salvar_chaves_api)
        btn_save_api.pack(anchor="e")

    def _abrir_modal_justificativa(self):
        if hasattr(self, "_select_tab"):
            self._select_tab("justificativa")

    def _setup_justificativa_ui(self):
        if getattr(self, "_justificativa_ui_built", False):
            self._sincronizar_dias_justificativa()
            return
        self._justificativa_ui_built = True

        top = self.frame_justificativa

        header_card = tk.Frame(
            top,
            bg=self.surface_color,
            padx=20,
            pady=14,
            highlightbackground=self.border_color,
            highlightthickness=1,
            bd=0
        )
        header_card.pack(fill="x", pady=(0, 12))

        tk.Label(
            header_card,
            text="📝 Central de Justificativas de Ponto para RH",
            bg=self.surface_color,
            fg=self.text_color,
            font=("Segoe UI", 12, "bold")
        ).pack(anchor="w")

        tk.Label(
            header_card,
            text="Marque os dias desejados, edite o ajuste proposto e o motivo. Os documentos serão gerados/enviados agrupados por semana.",
            bg=self.surface_color,
            fg=self.muted_text_color,
            font=("Segoe UI", 9)
        ).pack(anchor="w", pady=(2, 0))

        # Barra de Controles da Justificativa
        ctrl_card = tk.Frame(top, bg=self.surface_color, padx=16, pady=12, highlightbackground=self.border_color, highlightthickness=1)
        ctrl_card.pack(fill="x", pady=(0, 12))

        btn_sel_all = ttk.Button(ctrl_card, text="☑️ Marcar Todos", command=lambda: self._toggle_all_just_days(True))
        btn_sel_all.pack(side="left", padx=(0, 6))

        btn_unsel_all = ttk.Button(ctrl_card, text="☐ Desmarcar Todos", command=lambda: self._toggle_all_just_days(False))
        btn_unsel_all.pack(side="left", padx=(0, 15))

        self.lbl_just_count = ttk.Label(ctrl_card, text="0 dia(s) selecionado(s)", font=("Segoe UI", 9, "bold"))
        self.lbl_just_count.pack(side="left")

        btn_populate = ttk.Button(ctrl_card, text="🔄 Sincronizar com Espelho de Ponto", command=self._sincronizar_dias_justificativa)
        btn_populate.pack(side="right")

        # Tabela de Dias para Justificativa
        table_card = tk.Frame(top, bg=self.surface_color, padx=14, pady=12, highlightbackground=self.border_color, highlightthickness=1)
        table_card.pack(fill="both", expand=True, pady=(0, 12))

        cols_j = ("sel", "data", "dia_sem", "batidas_orig", "batidas_prop", "motivo")
        self.tree_just = ttk.Treeview(table_card, columns=cols_j, show="headings", selectmode="browse")
        self.tree_just.heading("sel", text="Sel")
        self.tree_just.heading("data", text="Data")
        self.tree_just.heading("dia_sem", text="Dia da Semana")
        self.tree_just.heading("batidas_orig", text="Batidas Registradas")
        self.tree_just.heading("batidas_prop", text="Ajuste Proposto (Horários)")
        self.tree_just.heading("motivo", text="Motivo / Justificativa")

        self.tree_just.column("sel", width=50, anchor="center")
        self.tree_just.column("data", width=100, anchor="center")
        self.tree_just.column("dia_sem", width=130, anchor="w")
        self.tree_just.column("batidas_orig", width=180, anchor="w")
        self.tree_just.column("batidas_prop", width=220, anchor="w")
        self.tree_just.column("motivo", width=260, anchor="w")

        vsb_j = ttk.Scrollbar(table_card, orient="vertical", command=self.tree_just.yview)
        self.tree_just.configure(yscrollcommand=vsb_j.set)

        self.tree_just.pack(side="left", fill="both", expand=True)
        vsb_j.pack(side="right", fill="y")

        def on_tree_just_click(event):
            region = self.tree_just.identify_region(event.x, event.y)
            item_id = self.tree_just.identify_row(event.y)
            column = self.tree_just.identify_column(event.x)
            if not item_id:
                return

            if column == "#1" or region == "cell":
                values = list(self.tree_just.item(item_id, "values"))
                curr_sel = values[0]
                new_sel = "[☐]" if curr_sel == "[☑]" else "[☑]"
                values[0] = new_sel
                self.tree_just.item(item_id, values=values)
                self._atualizar_contador_just_dias()

        def on_tree_just_dblclick(event):
            item_id = self.tree_just.identify_row(event.y)
            if not item_id:
                return
            values = list(self.tree_just.item(item_id, "values"))

            ed_top = tk.Toplevel(self)
            ed_top.title(f"Editar Justificativa - {values[1]}")
            ed_top.geometry("480x280")
            ed_top.transient(self)
            ed_top.grab_set()

            ttk.Label(ed_top, text=f"Data: {values[1]} ({values[2]})", font=("Segoe UI", 10, "bold")).pack(anchor="w", padx=15, pady=(15, 8))
            ttk.Label(ed_top, text="Batidas Propostas (ex: 08:00 12:00 13:00 17:00):").pack(anchor="w", padx=15)

            ent_prop = ttk.Entry(ed_top, width=48)
            ent_prop.insert(0, values[4])
            ent_prop.pack(anchor="w", padx=15, pady=(3, 10))

            ttk.Label(ed_top, text="Motivo / Justificativa:").pack(anchor="w", padx=15)
            ent_mot = ttk.Entry(ed_top, width=48)
            ent_mot.insert(0, values[5])
            ent_mot.pack(anchor="w", padx=15, pady=(3, 15))

            def salvar_edit():
                values[0] = "[☑]"
                values[4] = ent_prop.get().strip()
                values[5] = ent_mot.get().strip()
                self.tree_just.item(item_id, values=values)
                self._atualizar_contador_just_dias()
                ed_top.destroy()

            ttk.Button(ed_top, text="💾 Salvar Alterações", command=salvar_edit).pack(anchor="e", padx=15)

        self.tree_just.bind("<Button-1>", on_tree_just_click)
        self.tree_just.bind("<Double-1>", on_tree_just_dblclick)

        # Rodapé & Ações
        bot_card = tk.Frame(top, bg=self.surface_color, padx=16, pady=12, highlightbackground=self.border_color, highlightthickness=1)
        bot_card.pack(fill="x")

        ttk.Label(bot_card, text="Justificativa Geral / Observações:", bg=self.surface_color, font=("Segoe UI", 9, "bold")).pack(anchor="w", pady=(0, 4))
        self.txt_just_obs = tk.Text(bot_card, height=3, font=("Segoe UI", 9))
        self.txt_just_obs.pack(fill="x", pady=(0, 10))

        self.lbl_just_status = ttk.Label(bot_card, text="", font=("Segoe UI", 9, "bold"))
        self.lbl_just_status.pack(anchor="w", pady=(0, 6))

        btn_bar = ttk.Frame(bot_card)
        btn_bar.pack(fill="x")

        btn_pdf = ttk.Button(btn_bar, text="👁️ Gerar PDF(s) por Semana", command=self._gerar_pdfs_semanais)
        btn_pdf.pack(side="left", padx=(0, 10))

        btn_env_autentique = ttk.Button(btn_bar, text="🚀 Enviar via Autentique (por Semana)", command=self._enviar_autentique_semanal)
        btn_env_autentique.pack(side="left")

        self._sincronizar_dias_justificativa()

    def _sincronizar_dias_justificativa(self):
        if not hasattr(self, "tree_just"):
            return
        self.tree_just.delete(*self.tree_just.get_children())
        if not hasattr(self, "tree"):
            return

        for item_id in self.tree.get_children():
            vals = self.tree.item(item_id, "values")
            if not vals:
                continue
            # cols: ["Data", "E1", "S1", "E2", "S2", "E3", "S3", "Total", "Saldo", "Obs"]
            data_str = vals[0]
            punches = [v for v in vals[1:7] if v and v != "--:--"]
            batidas_orig = " ".join(punches) if punches else "Sem batidas"
            obs = vals[9] if len(vals) > 9 else ""

            try:
                dt = datetime.strptime(data_str, "%d/%m/%Y")
                dias_semana_br = ["SEGUNDA FEIRA", "TERÇA FEIRA", "QUARTA FEIRA", "QUINTA FEIRA", "SEXTA FEIRA", "SÁBADO", "DOMINGO"]
                dia_sem = dias_semana_br[dt.weekday()]
            except Exception:
                dia_sem = "-"

            is_checked = "[☑]" if (not punches or "-" in vals[8] or obs) else "[☐]"

            self.tree_just.insert("", "end", iid=item_id, values=(is_checked, data_str, dia_sem, batidas_orig, batidas_orig if punches else "08:00 12:00 13:00 17:00", obs or "Ajuste de ponto"))

        self._atualizar_contador_just_dias()

    def _toggle_all_just_days(self, mark: bool):
        if not hasattr(self, "tree_just"):
            return
        val_mark = "[☑]" if mark else "[☐]"
        for item_id in self.tree_just.get_children():
            vals = list(self.tree_just.item(item_id, "values"))
            vals[0] = val_mark
            self.tree_just.item(item_id, values=vals)
        self._atualizar_contador_just_dias()

    def _atualizar_contador_just_dias(self):
        if not hasattr(self, "tree_just") or not hasattr(self, "lbl_just_count"):
            return
        cnt = sum(1 for item_id in self.tree_just.get_children() if self.tree_just.item(item_id, "values")[0] == "[☑]")
        self.lbl_just_count.config(text=f"{cnt} dia(s) selecionado(s)")

    def _obter_dias_justificativa_agrupados_por_semana(self):
        if not hasattr(self, "tree_just"):
            return {}

        selected_items = []
        for item_id in self.tree_just.get_children():
            vals = self.tree_just.item(item_id, "values")
            if vals and vals[0] == "[☑]":
                selected_items.append({
                    "data": vals[1],
                    "dia_semana": vals[2],
                    "batidas_orig": vals[3],
                    "batidas_prop": vals[4],
                    "motivo": vals[5]
                })

        weeks_map = {}
        for item in selected_items:
            try:
                dt = datetime.strptime(item["data"], "%d/%m/%Y")
                iso_year, iso_week, _ = dt.isocalendar()
                key = (iso_year, iso_week)
                if key not in weeks_map:
                    weeks_map[key] = []

                pts = [p for p in item["batidas_prop"].split() if ":" in p]
                e1 = pts[0] if len(pts) > 0 else ""
                s1 = pts[1] if len(pts) > 1 else ""
                e2 = pts[2] if len(pts) > 2 else ""
                s2 = pts[3] if len(pts) > 3 else ""

                weeks_map[key].append({
                    "data": item["data"],
                    "dia_semana": item["dia_semana"],
                    "e1": e1, "s1": s1, "e2": e2, "s2": s2,
                    "motivo": item["motivo"]
                })
            except Exception as ex_dt:
                logging.error(f"Erro ao agrupar data {item.get('data')}: {ex_dt}")

        return weeks_map

    def _gerar_pdfs_semanais(self):
        weeks_map = self._obter_dias_justificativa_agrupados_por_semana()
        if not weeks_map:
            messagebox.showwarning("Aviso", "Selecione ao menos 1 dia para gerar a justificativa.")
            return

        def safe_keyring_get(key, default=""):
            try:
                return keyring.get_password("DixiPontoApp", key) or default
            except Exception:
                return default

        colab_nome = safe_keyring_get("colaborador_nome", "Colaborador")
        colab_cargo = safe_keyring_get("colaborador_cargo", "")
        gestor_nome = safe_keyring_get("gestor_nome", "Gestor Imediato")
        rh_nome = safe_keyring_get("rh_nome", "Recursos Humanos")
        obs_geral = self.txt_just_obs.get("1.0", "end").strip() if hasattr(self, "txt_just_obs") else ""

        pdf_paths = []
        output_dir = os.path.join(os.path.expanduser("~"), "Downloads")
        os.makedirs(output_dir, exist_ok=True)

        for (year, week), itens in weeks_map.items():
            pdf_name = f"Justificativa_Ponto_Semana_{week}_{year}.pdf"
            output_pdf = os.path.join(output_dir, pdf_name)
            gerar_pdf_justificativa(
                colaborador_nome=colab_nome,
                colaborador_funcao=colab_cargo,
                mes_competencia_str=itens[0]["data"],
                data_solicitacao=datetime.now().strftime("%d/%m/%Y"),
                justificativa_geral=obs_geral,
                gestor_nome=gestor_nome,
                rh_nome=rh_nome,
                itens_ponto=itens,
                output_pdf_path=output_pdf
            )
            pdf_paths.append(output_pdf)

        messagebox.showinfo("PDFs Gerados", f"Foi(ram) gerado(s) {len(pdf_paths)} PDF(s) por semana em:\n{output_dir}")
        for p in pdf_paths:
            try:
                os.startfile(p)
            except Exception:
                pass

    def _enviar_autentique_semanal(self):
        weeks_map = self._obter_dias_justificativa_agrupados_por_semana()
        if not weeks_map:
            messagebox.showwarning("Aviso", "Selecione ao menos 1 dia para enviar a justificativa.")
            return

        def safe_keyring_get(key, default=""):
            try:
                return keyring.get_password("DixiPontoApp", key) or default
            except Exception:
                return default

        token = safe_keyring_get("autentique_token")
        if not token:
            messagebox.showwarning("Token Ausente", "Token do Autentique não configurado. Por favor, cadastre o Token em '⚙️ Configurações'.")
            return

        colab_email = safe_keyring_get("colaborador_email", "")
        gestor_email = safe_keyring_get("gestor_email", "")
        rh_email = safe_keyring_get("rh_email", "")

        if not colab_email or not gestor_email:
            messagebox.showwarning("E-mails Ausentes", "Por favor, cadastre os e-mails do Colaborador e do Gestor no botão '⚙️ Configurações'.")
            return

        colab_nome = safe_keyring_get("colaborador_nome", "Colaborador")
        colab_cargo = safe_keyring_get("colaborador_cargo", "")
        gestor_nome = safe_keyring_get("gestor_nome", "Gestor Imediato")
        rh_nome = safe_keyring_get("rh_nome", "Recursos Humanos")
        obs_geral = self.txt_just_obs.get("1.0", "end").strip() if hasattr(self, "txt_just_obs") else ""

        self.lbl_just_status.config(text="⏳ Gerando PDFs e enviando semanas para o Autentique...", foreground="#0284c7")

        def run_async():
            try:
                output_dir = os.path.join(os.path.expanduser("~"), "Downloads")
                os.makedirs(output_dir, exist_ok=True)
                count_sent = 0

                for (year, week), itens in weeks_map.items():
                    pdf_name = f"Justificativa_Ponto_Semana_{week}_{year}.pdf"
                    output_pdf = os.path.join(output_dir, pdf_name)
                    gerar_pdf_justificativa(
                        colaborador_nome=colab_nome,
                        colaborador_funcao=colab_cargo,
                        mes_competencia_str=itens[0]["data"],
                        data_solicitacao=datetime.now().strftime("%d/%m/%Y"),
                        justificativa_geral=obs_geral,
                        gestor_nome=gestor_nome,
                        rh_nome=rh_nome,
                        itens_ponto=itens,
                        output_pdf_path=output_pdf
                    )

                    signers = [
                        {"email": colab_email, "action": "SIGN"},
                        {"email": gestor_email, "action": "SIGN"}
                    ]
                    if rh_email:
                        signers.append({"email": rh_email, "action": "SIGN"})

                    enviar_justificativa_autentique(
                        token=token,
                        caminho_pdf=output_pdf,
                        nome_documento=f"Justificativa Ponto - {colab_nome} - Sem {week}/{year}",
                        lista_signatarios=signers
                    )
                    count_sent += 1

                def update_success():
                    self.lbl_just_status.config(text=f"✅ {count_sent} documento(s) semanal(is) enviado(s) ao Autentique!", foreground="#16a34a")
                    messagebox.showinfo("Sucesso", f"{count_sent} documento(s) agrupado(s) por semana foi(ram) enviado(s) ao Autentique!")
                    self.after(1500, lambda: self._select_tab("historico"))

                self.after(0, update_success)

            except Exception as ex:
                err_msg = str(ex)
                def update_err():
                    self.lbl_just_status.config(text=f"❌ Erro ao enviar: {err_msg}", foreground="#dc2626")
                self.after(0, update_err)


    def _setup_historico_ui(self):


        # Card Superior: Cabeçalho & Ações
        header_card = tk.Frame(
            self.frame_historico,
            bg=self.surface_color,
            padx=20,
            pady=16,
            highlightbackground=self.border_color,
            highlightthickness=1,
            bd=0
        )
        header_card.pack(fill="x", pady=(0, 14))

        top_row = tk.Frame(header_card, bg=self.surface_color)
        top_row.pack(fill="x")

        title_frame = tk.Frame(top_row, bg=self.surface_color)
        title_frame.pack(side="left", fill="x", expand=True)


        tk.Label(
            title_frame,
            text="📜 Histórico de Justificativas (Autentique)",
            bg=self.surface_color,
            fg=self.text_color,
            font=("Segoe UI", 12, "bold")
        ).pack(anchor="w")

        tk.Label(
            title_frame,
            text="Acompanhe em tempo real quem já assinou e quem ainda falta assinar nos documentos enviados ao RH.",
            bg=self.surface_color,
            fg=self.muted_text_color,
            font=("Segoe UI", 9)
        ).pack(anchor="w", pady=(2, 0))

        self.btn_refresh_hist = ttk.Button(
            top_row,
            text="🔄 Atualizar Histórico",
            command=self._carregar_historico_autentique_ui
        )
        self.btn_refresh_hist.pack(side="right")

        self.lbl_hist_status = ttk.Label(header_card, text="", font=("Segoe UI", 9, "italic"))
        self.lbl_hist_status.pack(anchor="w", pady=(8, 0))

        # PanedWindow Container Card
        paned_card = tk.Frame(
            self.frame_historico,
            bg=self.surface_color,
            padx=16,
            pady=16,
            highlightbackground=self.border_color,
            highlightthickness=1,
            bd=0
        )
        paned_card.pack(fill="both", expand=True)

        paned = ttk.PanedWindow(paned_card, orient="vertical")
        paned.pack(fill="both", expand=True)

        # 1. Tabela Superior: Documentos
        frame_docs = ttk.LabelFrame(paned_card, text=" 📂 Documentos Enviados ao Autentique ", padding=8)
        paned.add(frame_docs, weight=1)

        cols_docs = ("name", "date", "status", "sigs_count")
        self.tree_docs_hist = ttk.Treeview(frame_docs, columns=cols_docs, show="headings", selectmode="browse")
        self.tree_docs_hist.heading("name", text="Nome do Documento")
        self.tree_docs_hist.heading("date", text="Data Envio")
        self.tree_docs_hist.heading("status", text="Status Geral")
        self.tree_docs_hist.heading("sigs_count", text="Assinaturas")

        self.tree_docs_hist.column("name", width=380, anchor="w")
        self.tree_docs_hist.column("date", width=140, anchor="center")
        self.tree_docs_hist.column("status", width=160, anchor="center")
        self.tree_docs_hist.column("sigs_count", width=110, anchor="center")

        vsb_docs = ttk.Scrollbar(frame_docs, orient="vertical", command=self.tree_docs_hist.yview)
        self.tree_docs_hist.configure(yscrollcommand=vsb_docs.set)

        self.tree_docs_hist.pack(side="left", fill="both", expand=True)
        vsb_docs.pack(side="right", fill="y")

        # 2. Tabela Inferior: Signatários
        frame_sigs = ttk.LabelFrame(paned_card, text=" 🔍 Signatários e Status do Documento Selecionado ", padding=8)
        paned.add(frame_sigs, weight=1)

        cols_sigs = ("name", "email", "role", "status", "date")
        self.tree_sigs_hist = ttk.Treeview(frame_sigs, columns=cols_sigs, show="headings", selectmode="browse")
        self.tree_sigs_hist.heading("name", text="Nome")
        self.tree_sigs_hist.heading("email", text="E-mail")
        self.tree_sigs_hist.heading("role", text="Ação / Papel")
        self.tree_sigs_hist.heading("status", text="Status de Assinatura")
        self.tree_sigs_hist.heading("date", text="Data do Evento")

        self.tree_sigs_hist.column("name", width=180, anchor="w")
        self.tree_sigs_hist.column("email", width=220, anchor="w")
        self.tree_sigs_hist.column("role", width=100, anchor="center")
        self.tree_sigs_hist.column("status", width=190, anchor="center")
        self.tree_sigs_hist.column("date", width=140, anchor="center")

        vsb_sigs = ttk.Scrollbar(frame_sigs, orient="vertical", command=self.tree_sigs_hist.yview)
        self.tree_sigs_hist.configure(yscrollcommand=vsb_sigs.set)

        self.tree_sigs_hist.pack(side="left", fill="both", expand=True)
        vsb_sigs.pack(side="right", fill="y")

        # Barra Inferior com Info & Link
        bottom_actions = tk.Frame(paned_card, bg=self.surface_color, pady=8)
        bottom_actions.pack(fill="x")

        self.lbl_hist_info = ttk.Label(
            bottom_actions,
            text="Selecione um documento na tabela acima para ver quem já assinou e quem falta assinar.",
            font=("Segoe UI", 9)
        )
        self.lbl_hist_info.pack(side="left")

        def abrir_link_assinatura():
            sel = self.tree_sigs_hist.selection()
            if not sel:
                messagebox.showinfo("Aviso", "Selecione um signatário na tabela inferior.")
                return
            item = self.tree_sigs_hist.item(sel[0])
            link = item.get("tags", [""])[0] if item.get("tags") else ""
            if link and link.startswith("http"):
                import webbrowser
                webbrowser.open(link)
            else:
                messagebox.showinfo("Aviso", "Nenhum link de assinatura disponível para este signatário.")

        btn_open_link = ttk.Button(bottom_actions, text="🌐 Abrir Link de Assinatura", command=abrir_link_assinatura)
        btn_open_link.pack(side="right")

        self.tree_docs_hist.bind("<<TreeviewSelect>>", self._on_hist_doc_select)
        self.hist_docs_cache = {}

    def _on_hist_doc_select(self, event):
        selected = self.tree_docs_hist.selection()
        self.tree_sigs_hist.delete(*self.tree_sigs_hist.get_children())

        if not selected:
            self.lbl_hist_info.config(text="Selecione um documento acima para ver os signatários.")
            return

        doc_id = selected[0]
        doc_data = self.hist_docs_cache.get(doc_id)
        if not doc_data:
            return

        signatures = doc_data.get("signatures") or []
        signed_cnt = 0

        def format_iso_date(iso_str):
            if not iso_str:
                return "-"
            try:
                dt = datetime.fromisoformat(iso_str.replace("Z", "+00:00"))
                return dt.strftime("%d/%m/%Y %H:%M")
            except Exception:
                return iso_str[:19].replace("T", " ")

        for i, sig in enumerate(signatures):
            sig_name = sig.get("name") or (sig.get("user") or {}).get("name") or "-"
            sig_email = sig.get("email") or (sig.get("user") or {}).get("email") or "-"
            sig_action = (sig.get("action") or {}).get("name", "SIGN")

            signed_info = sig.get("signed")
            rejected_info = sig.get("rejected")
            viewed_info = sig.get("viewed")
            short_link = (sig.get("link") or {}).get("short_link", "")

            if signed_info and signed_info.get("created_at"):
                st_text = "✅ ASSINOU"
                st_date = format_iso_date(signed_info.get("created_at"))
                signed_cnt += 1
            elif rejected_info and rejected_info.get("created_at"):
                st_text = "❌ RECUSOU"
                st_date = format_iso_date(rejected_info.get("created_at"))
            elif viewed_info and viewed_info.get("created_at"):
                st_text = "👁️ Visualizou (Pendente)"
                st_date = format_iso_date(viewed_info.get("created_at"))
            else:
                st_text = "⏳ Falta Assinar"
                st_date = "-"

            item_id = f"sig_{i}"
            self.tree_sigs_hist.insert("", "end", iid=item_id, values=(sig_name, sig_email, sig_action, st_text, st_date), tags=(short_link,))

        self.lbl_hist_info.config(text=f"Documento '{doc_data.get('name')}': {signed_cnt} de {len(signatures)} assinatura(s) concluída(s).")

    def _carregar_historico_autentique_ui(self):
        def safe_keyring_get(key, default=""):
            try:
                return keyring.get_password("DixiPontoApp", key) or default
            except Exception:
                return default

        tok = safe_keyring_get("autentique_token")
        if not tok:
            self.lbl_hist_status.config(text="⚠️ Token do Autentique não configurado. Cadastre o token em 'Justificar Ponto para RH'.", foreground="#d97706")
            return

        self.btn_refresh_hist.config(state="disabled")
        self.lbl_hist_status.config(text="⏳ Buscando documentos no Autentique...", foreground="#0284c7")
        self.tree_docs_hist.delete(*self.tree_docs_hist.get_children())
        self.tree_sigs_hist.delete(*self.tree_sigs_hist.get_children())
        self.hist_docs_cache.clear()

        def format_iso_date(iso_str):
            if not iso_str:
                return "-"
            try:
                dt = datetime.fromisoformat(iso_str.replace("Z", "+00:00"))
                return dt.strftime("%d/%m/%Y %H:%M")
            except Exception:
                return iso_str[:19].replace("T", " ")

        def fetch():
            try:
                res = listar_documentos_autentique(tok, page=1, limit=60)
                doc_list = res.get("data") or []

                def update_ui():
                    self.lbl_hist_status.config(text=f"✅ {len(doc_list)} documento(s) encontrado(s).", foreground="#16a34a")
                    self.btn_refresh_hist.config(state="normal")
                    if not doc_list:
                        self.lbl_hist_info.config(text="Nenhum documento encontrado na conta do Autentique.")
                        return

                    for doc in doc_list:
                        doc_id = doc.get("id")
                        self.hist_docs_cache[doc_id] = doc
                        doc_name = doc.get("name") or "Sem nome"
                        created_at = format_iso_date(doc.get("created_at"))

                        signatures = doc.get("signatures") or []
                        total_sigs = len(signatures)
                        signed_count = sum(1 for s in signatures if s.get("signed") and s.get("signed").get("created_at"))
                        rejected_count = sum(1 for s in signatures if s.get("rejected") and s.get("rejected").get("created_at"))

                        if total_sigs > 0 and signed_count == total_sigs:
                            status_geral = "🟢 Todos Assinaram"
                        elif rejected_count > 0:
                            status_geral = "🔴 Recusado"
                        else:
                            status_geral = "🟡 Pendente"

                        sigs_str = f"{signed_count}/{total_sigs} assinado(s)"
                        self.tree_docs_hist.insert("", "end", iid=doc_id, values=(doc_name, created_at, status_geral, sigs_str))

                self.after(0, update_ui)

            except Exception as ex:
                err_msg = str(ex)
                def update_err():
                    self.lbl_hist_status.config(text=f"❌ Erro ao buscar no Autentique: {err_msg}", foreground="#dc2626")
                    self.btn_refresh_hist.config(state="normal")
                self.after(0, update_err)

        threading.Thread(target=fetch, daemon=True).start()

if __name__ == "__main__":
    AppPonto().mainloop()





