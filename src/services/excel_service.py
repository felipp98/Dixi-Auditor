"""
Serviço de geração e formatação de planilhas de auditoria em Excel (OpenPyXL).
"""
import logging
from datetime import datetime
from typing import List

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.core.models import MarcacaoDia
from src.utils.formatters import format_time_seconds

logger = logging.getLogger(__name__)

class ExcelService:
    """Exporta relatórios consolidados em Excel com formatação condicional e cabeçalhos auditáveis."""

    @classmethod
    def generate(cls, data: List[MarcacaoDia], output_path: str, ignore_today: bool = True) -> str:
        """Gera a planilha Excel completa com dados de ponto e estilos visuais."""
        rows = []
        sum_saldo = 0
        today_str = datetime.now().strftime("%d/%m/%Y")

        max_horarios = max([len(m.horarios) for m in data]) if data else 0
        max_cols = max(6, max_horarios)
        if max_cols % 2 != 0:
            max_cols += 1

        for m in data:
            is_today = (m.data_formatada == today_str) and ignore_today
            if not is_today:
                sum_saldo += m.saldo_segundos

            punches = (m.horarios + [""] * max_cols)[:max_cols]
            saldo_str = "00:00" if is_today else format_time_seconds(m.saldo_segundos, show_sign=True)
            obs_str = m.obs if m.obs else ("EM ANDAMENTO" if is_today else ("FALTA BATIDA" if m.is_pendencia else ""))

            row = [m.data_formatada] + punches + [
                format_time_seconds(m.segundos_trabalhados),
                saldo_str,
                obs_str
            ]
            rows.append(row)

        headers = ["Data"]
        for i in range(1, (max_cols // 2) + 1):
            headers.extend([f"E{i}", f"S{i}"])
        headers.extend(["Total", "Saldo", "Obs"])

        df = pd.DataFrame(rows, columns=headers)
        df.to_excel(output_path, index=False)

        cls._apply_excel_formatting(output_path, sum_saldo, max_cols)
        return output_path

    @classmethod
    def _apply_excel_formatting(cls, path: str, total_saldo_segundos: int, max_cols: int):
        """Aplica cores condicionais, fontes e linha de saldo acumulado no Excel."""
        wb = load_workbook(path)
        ws = wb.active
        last_row = ws.max_row + 1

        col_saldo_letter = get_column_letter(3 + max_cols)
        col_obs_letter = get_column_letter(4 + max_cols)

        ws[f"A{last_row}"] = "SALDO ACUMULADO"
        ws[f"{col_saldo_letter}{last_row}"] = format_time_seconds(total_saldo_segundos, show_sign=True)

        fill_total = PatternFill(start_color="DDEBF7", fill_type="solid")
        fill_red = PatternFill(start_color="FFC7CE", fill_type="solid")
        fill_green = PatternFill(start_color="C6EFCE", fill_type="solid")
        fill_yellow = PatternFill(start_color="FFF2CC", fill_type="solid")
        fill_olive = PatternFill(start_color="E2EFDA", fill_type="solid")
        bold_font = Font(bold=True)

        # Estiliza a linha de total
        for cell in ws[last_row]:
            cell.fill = fill_total
            cell.font = bold_font

        # Formatação condicional nas linhas de dados
        for row in range(2, ws.max_row):
            saldo_cell = ws[f"{col_saldo_letter}{row}"]
            obs_cell = ws[f"{col_obs_letter}{row}"]
            val = str(saldo_cell.value)

            if "+" in val:
                saldo_cell.fill = fill_green
            elif "-" in val and val != "00:00":
                saldo_cell.fill = fill_red

            if obs_cell.value == "FALTA BATIDA":
                obs_cell.fill = fill_red
            elif obs_cell.value == "EM ANDAMENTO":
                obs_cell.fill = fill_olive
            elif obs_cell.value and any(k in str(obs_cell.value) for k in ["IA", "Ajust", "Abon"]):
                obs_cell.fill = fill_yellow

        # Ajuste de largura das colunas
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

        wb.save(path)
